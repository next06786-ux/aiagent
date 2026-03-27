import os
import time
import math
import torch
import torch.nn as nn
import transformers
from bigptq import BRAGPTQ
from binary import (
    Binarization,
    quantize_to_2bit_perchannel,
    quantize_to_3bit_perchannel,
    quantize_to_4bit_perchannel,
    quantize_to_8bit_perchannel,
)
from modelutils import find_layers
from loguru import logger
from transformers.models.llama import modeling_llama
from transformers import AutoConfig
_original_rotary_forward = modeling_llama.LlamaRotaryEmbedding.forward

QWEN2_AVAILABLE = False
QWEN3_AVAILABLE = False
_original_qwen2_rotary_forward = None
_original_qwen3_rotary_forward = None

try:
    from transformers.models.qwen2 import modeling_qwen2
    _original_qwen2_rotary_forward = modeling_qwen2.Qwen2RotaryEmbedding.forward
    QWEN2_AVAILABLE = True
except Exception:
    modeling_qwen2 = None

try:
    from transformers.models.qwen3 import modeling_qwen3
    _original_qwen3_rotary_forward = modeling_qwen3.Qwen3RotaryEmbedding.forward
    QWEN3_AVAILABLE = True
except Exception:
    modeling_qwen3 = None


def _patched_rotary_forward(self, x, *args, **kwargs):

    target_device = x.device if isinstance(x, torch.Tensor) else None
    
    # 检查 position_ids (可能在 args[0] 或 kwargs['position_ids'])
    position_ids = None
    if len(args) > 0:
        position_ids = args[0]
    if 'position_ids' in kwargs:
        position_ids = kwargs['position_ids']
    
    # ⭐ 如果 position_ids 为 None，为新版本 transformers 自动生成
    if position_ids is None:
        # --- 修改开始 ---
        # 修复：使用 shape[-2] 获取 seq_len，避免将 num_heads 误判为 seq_len
        seq_len = x.shape[-2] 
        batch_size = x.shape[0]
        # --- 修改结束 ---
        
        position_ids = torch.arange(seq_len, dtype=torch.long, device=x.device)
        position_ids = position_ids.unsqueeze(0).expand(batch_size, -1)
        
        # 更新 args 或 kwargs
        if len(args) > 0:
            args = (position_ids,) + args[1:]
        else:
            kwargs['position_ids'] = position_ids
    
    # 确保设备一致性
    for arg in args:
        if isinstance(arg, torch.Tensor):
            target_device = arg.device
            break
    
    if 'position_ids' in kwargs and isinstance(kwargs['position_ids'], torch.Tensor):
        target_device = kwargs['position_ids'].device
    
    if target_device is not None and hasattr(self, 'inv_freq'):
        if self.inv_freq.device != target_device:
            self.inv_freq = self.inv_freq.to(target_device)
    
    # 根据模型类型调用原始函数
    class_name = self.__class__.__name__ if hasattr(self, "__class__") else ""
    if QWEN3_AVAILABLE and "Qwen3" in class_name:
        return _original_qwen3_rotary_forward(self, x, *args, **kwargs)
    if QWEN2_AVAILABLE and "Qwen2" in class_name:
        return _original_qwen2_rotary_forward(self, x, *args, **kwargs)
    return _original_rotary_forward(self, x, *args, **kwargs)

modeling_llama.LlamaRotaryEmbedding.forward = _patched_rotary_forward
print("Successfully patched LlamaRotaryEmbedding for device consistency.")

if QWEN2_AVAILABLE:
    modeling_qwen2.Qwen2RotaryEmbedding.forward = _patched_rotary_forward
    print("Successfully patched Qwen2RotaryEmbedding for device consistency.")
if QWEN3_AVAILABLE:
    modeling_qwen3.Qwen3RotaryEmbedding.forward = _patched_rotary_forward
    print("Successfully patched Qwen3RotaryEmbedding for device consistency.")


def group_layers_by_block(layer_sensitivities):

    blocks = {}
    for layer_key in layer_sensitivities.keys():
        block_id = int(layer_key.split('.')[0])
        if block_id not in blocks:
            blocks[block_id] = []
        blocks[block_id].append(layer_key)
    
    # 对每个 Block 内的层进行排序（保持一致性）
    for block_id in blocks:
        blocks[block_id].sort()
    
    return blocks


def allocate_block_budgets_by_entropy(blocks, layer_activation_entropies, 
                                       min_avg_rank=128, max_avg_rank=512):

    # 1. 计算每个block的平均激活熵
    block_avg_entropies = {}
    for block_id, layers in blocks.items():
        entropies = []
        for layer in layers:
            entropy = layer_activation_entropies.get(layer, 1.0)
            if entropy <= 0:
                logger.warning(f"Block {block_id} layer {layer}: invalid entropy={entropy:.6f}, using 1.0")
                entropy = 1.0
            entropies.append(entropy)
        
        block_avg_entropies[block_id] = sum(entropies) / len(entropies) if entropies else 1.0
    
    # 2. 找到全局最小和最大激活熵
    min_entropy = min(block_avg_entropies.values())
    max_entropy = max(block_avg_entropies.values())
    
    logger.info("=" * 80)
    logger.info("Block-level budget allocation (based on activation entropy)")
    logger.info(f"Min block avg entropy: {min_entropy:.6f}")
    logger.info(f"Max block avg entropy: {max_entropy:.6f}")
    logger.info(f"Rank range: [{min_avg_rank}, {max_avg_rank}]")
    logger.info("=" * 80)
    
    # 3. 线性映射：entropy -> avg_rank
    block_budgets = {}
    block_avg_ranks = {}
    total_budget = 0
    
    for block_id, avg_entropy in sorted(block_avg_entropies.items()):
        num_layers = len(blocks[block_id])
        
        # 线性映射
        if max_entropy > min_entropy:
            # 标准化到 [0, 1]
            normalized = (avg_entropy - min_entropy) / (max_entropy - min_entropy)
            # 映射到 [min_avg_rank, max_avg_rank]
            avg_rank = min_avg_rank + normalized * (max_avg_rank - min_avg_rank)
        else:
            # 所有block的熵相同，使用中间值
            avg_rank = (min_avg_rank + max_avg_rank) / 2.0
        
        # 取整并计算总预算
        avg_rank = int(round(avg_rank))
        block_budget = num_layers * avg_rank
        
        block_budgets[block_id] = block_budget
        block_avg_ranks[block_id] = avg_rank
        total_budget += block_budget
        
        logger.info(f"Block {block_id}: avg_entropy={avg_entropy:.6f}, "
                   f"avg_rank={avg_rank}, num_layers={num_layers}, "
                   f"total_budget={block_budget}")
    
    logger.info("=" * 80)
    logger.info(f"Total budget across all blocks: {total_budget}")
    logger.info("=" * 80)
    
    return block_budgets, block_avg_ranks


def allocate_ranks_block_greedy(blocks, entropy_reduction_tables, 
                                 base_rank, step_size, layer_max_rank, avg_rank,
                                 layer_activation_norms=None, 
                                layer_activation_entropies=None,
                                 activation_entropy_beta=2.0, use_activation_entropy=False,
                                 layer_stable_ranks=None, block_budgets=None):

    from tqdm import tqdm
  
    if layer_activation_norms is None:
        layer_activation_norms = {}
        logger.warning("layer_activation_norms not provided, using uniform activation importance")
    
    if layer_stable_ranks is None:
        layer_stable_ranks = {}
    
    rank_allocations = {}
    
    for block_id, block_layers in tqdm(sorted(blocks.items()), desc="Block greedy allocation"):
        # 计算块内总预算
        num_layers = len(block_layers)
        
        # ⭐ 如果提供了block_budgets，则使用预先分配的预算；否则使用avg_rank
        if block_budgets is not None and block_id in block_budgets:
            block_budget = block_budgets[block_id]
            logger.info(f"Block {block_id}: using pre-allocated budget={block_budget} "
                       f"(avg_rank={block_budget/num_layers:.1f})")
        else:
            block_budget = num_layers * avg_rank
            if block_budgets is not None:
                logger.warning(f"Block {block_id}: budget not found in block_budgets, "
                             f"using default (num_layers={num_layers} × avg_rank={avg_rank})")
        
        # 预分配
        current_ranks = {layer: base_rank for layer in block_layers}
        free_budget = block_budget - base_rank * num_layers
        

        logger.info(f"Block {block_id}: {num_layers} layers, "
                   f"budget={block_budget}, free_budget={free_budget}, "
                  )
        
        logger.info(f"Block {block_id} initial layer states:")
        for layer_key in block_layers:            
            initial_rank = current_ranks[layer_key]
            has_entropy_table = layer_key in entropy_reduction_tables
            
            # 获取稳定秩（如果可用）
            stable_rank = layer_stable_ranks.get(layer_key, None) if layer_stable_ranks else None
            
            log_msg = (f"  {layer_key}: "
                      f"initial_rank={initial_rank}, "
                      f"has_entropy_table={has_entropy_table}")
            if stable_rank is not None:
                log_msg += f", stable_rank={stable_rank:.6f}"
            logger.info(log_msg)
            
            # 检查异常情况
            if not has_entropy_table:
                logger.warning(f"  ⚠️  {layer_key}: missing entropy_reduction_table!")
        
        # 贪心循环
        iteration = 0
        has_valid_candidate = False
        
        while free_budget >= step_size:
            best_layer = None
            best_weighted_entropy_reduction = -float('inf')
            has_valid_candidate = False
            
            for layer_key in block_layers:
                current_rank = current_ranks[layer_key]
                next_rank = current_rank + step_size
                
                # 检查是否超过单层最大rank
                if next_rank > layer_max_rank:
                    continue
    
                # ⭐ 计算熵减（用于方法2或日志对比）
                entropy_reduction = None
                if layer_key in entropy_reduction_tables:
                    entropy_table = entropy_reduction_tables[layer_key]
                    if current_rank in entropy_table and next_rank in entropy_table:
                        entropy_reduction = entropy_table[next_rank] - entropy_table[current_rank]
                
                # ⭐ 根据 use_activation_entropy 选择计算方法
                if use_activation_entropy:
                    # 方法2：基于稳定秩的熵减
                    # Score_l = ΔH_l × log10(stable_rank)
                    if entropy_reduction is None:
                        logger.warning(f"  ⚠️  {layer_key}: entropy_reduction not available, skipping.")
                        continue
                    
                    # 获取稳定秩
                    if layer_key in layer_stable_ranks:
                        stable_rank = layer_stable_ranks[layer_key]
                        if stable_rank <= 0:
                            logger.warning(f"  ⚠️  {layer_key}: invalid stable_rank={stable_rank:.6f} (<=0), using default 1.0")
                            stable_rank = 1.0
                    else:
                        stable_rank = 1.0
                        logger.warning(f"  ⚠️  {layer_key}: stable_rank not found, using default 1.0")
                    
                    # 计算 log10(stable_rank)
                    log_stable_rank = math.log10(stable_rank)
                    
                    # 计算加权分数
                    weighted_score = entropy_reduction * log_stable_rank
                    
                    # 获取激活范数用于日志（不用于计算 score）
                    avg_activation_norm = layer_activation_norms.get(layer_key, 1.0) if layer_activation_norms else 1.0
                    activation_entropy = layer_activation_entropies.get(layer_key, None) if layer_activation_entropies else None

                # 标记有有效候选
                has_valid_candidate = True
                
                # ⭐ 使用加权分数进行比较（根据方法选择）
                if weighted_score > best_weighted_entropy_reduction:
                    best_weighted_entropy_reduction = weighted_score
                    best_layer = layer_key
            
            # 检查是否有有效候选
            if not has_valid_candidate:
                logger.warning(f"Block {block_id} Iter {iteration + 1}: no valid candidates found! Breaking.")
                break
            
            if best_layer:
                current_ranks[best_layer] += step_size
                free_budget -= step_size
                iteration += 1
                
                # ⭐ 记录分配信息（重新计算以获取准确的值）
                best_current_rank = current_ranks[best_layer] - step_size
                best_next_rank = current_ranks[best_layer]
                
                # 根据方法计算相应的值
                if use_activation_entropy:
                    # 方法2：基于稳定秩的熵减
                    best_entropy_reduction = None
                    if best_layer in entropy_reduction_tables:
                        best_entropy_table = entropy_reduction_tables[best_layer]
                        if best_current_rank in best_entropy_table and best_next_rank in best_entropy_table:
                            best_entropy_reduction = best_entropy_table[best_next_rank] - best_entropy_table[best_current_rank]
                    
                    # 获取稳定秩
                    best_stable_rank = layer_stable_ranks.get(best_layer, 1.0) if layer_stable_ranks else 1.0
                    if best_stable_rank <= 0:
                        best_stable_rank = 1.0
                    best_log_stable_rank = math.log10(best_stable_rank)
                    
                    # 保留激活熵用于日志（如果可用）
                    best_activation_entropy = layer_activation_entropies.get(best_layer, None) if layer_activation_entropies else None
                    best_energy_gain = None
                    best_activation_norm = None
                    best_activation_norm_squared = None

                
                if use_activation_entropy:
                    # 方法2：基于稳定秩的熵减
                    log_msg = (f"Block {block_id} Iter {iteration}: ✅ Allocated +{step_size} to {best_layer}, "
                              f"rank: {best_current_rank} -> {best_next_rank}, "
                              f"entropy_reduction={best_entropy_reduction:.6f}, "
                              f"stable_rank={best_stable_rank:.6f}, "
                              f"log10(stable_rank)={best_log_stable_rank:.6f}, "
                              f"weighted_score={best_weighted_entropy_reduction:.6f}, "
                              f"remaining_budget={free_budget}")
                    if best_activation_entropy is not None:
                        log_msg += f", activation_entropy={best_activation_entropy:.6f} (for comparison)"
                else:
                    # 方法1：基于绝对能量增益
                    log_msg = (f"Block {block_id} Iter {iteration}: ✅ Allocated +{step_size} to {best_layer}, "
                              f"rank: {best_current_rank} -> {best_next_rank}, "
                              f"energy_gain={best_energy_gain:.6f}, "
                              f"activation_norm={best_activation_norm:.6f}, "
                              f"activation_norm_squared={best_activation_norm_squared:.6f}, "
                              f"weighted_score={best_weighted_entropy_reduction:.6f}, "
                              f"remaining_budget={free_budget}")
                    if best_entropy_reduction is not None:
                        log_msg += f", entropy_reduction={best_entropy_reduction:.6f} (for comparison)"
                    if best_activation_entropy is not None:
                        log_msg += f", activation_entropy={best_activation_entropy:.6f} (for comparison)"
                logger.info(log_msg)
            else:
                logger.warning(f"Block {block_id} Iter {iteration + 1}: no best_layer selected! Breaking.")
                break  # 所有层都达到上限
        
        # 记录分配结果
        for layer_key, rank in current_ranks.items():
            rank_allocations[layer_key] = rank
        
        logger.info(f"Block {block_id} final allocation summary:")
        total_rank_used = 0
        for layer_key in block_layers:
            final_rank = current_ranks[layer_key]
            total_rank_used += final_rank
            activation_norm = layer_activation_norms.get(layer_key, 1.0)
            if activation_norm <= 0:
                activation_norm = 1.0

            rank_increase = final_rank - base_rank
            
            stable_rank = layer_stable_ranks.get(layer_key, None) if layer_stable_ranks else None
            
            log_msg = (f"  {layer_key}: "
                    f"final_rank={final_rank} (base={base_rank}, +{rank_increase}), "
                    f"activation_norm={activation_norm:.6f}, "
                   )
            if stable_rank is not None:
                log_msg += f", stable_rank={stable_rank:.6f}"
            logger.info(log_msg)
        
        logger.info(f"Block {block_id} total rank used: {total_rank_used} / {block_budget} "
                   f"(budget_utilization={total_rank_used/block_budget*100:.2f}%)")
        
        if abs(total_rank_used - block_budget) > step_size:
            logger.warning(f"  ⚠️  Block {block_id}: rank usage mismatch! "
                         f"Used {total_rank_used} but budget was {block_budget}, "
                         f"difference={total_rank_used - block_budget}")
        
        logger.info(f"Block {block_id} final allocation: {dict(current_ranks)}")
    

    
    return rank_allocations


def allocate_ranks(layer_sensitivities, total_budget, min_rank=32, max_rank=1024,
                   block_greedy=False, entropy_reduction_tables=None, 
                   base_rank=128, step_size=32, avg_rank=256,
                   layer_activation_norms=None, 
                    layer_activation_entropies=None,
                   activation_entropy_beta=2.0, use_activation_entropy=False,
                   layer_stable_ranks=None, layer_max_rank=None):
    
    if block_greedy:
        # 块内贪心分配
        if entropy_reduction_tables is None:
            raise ValueError("entropy_reduction_tables is required for block_greedy mode")
        
        # 检查所有层是否都有熵减表
        missing_tables = []
        for layer_key in layer_sensitivities.keys():
            if layer_key not in entropy_reduction_tables:
                missing_tables.append(layer_key)
        
        if missing_tables:
            raise ValueError(f"Missing entropy_reduction_tables for {len(missing_tables)} layers: {missing_tables[:10]}...")
        
        blocks = group_layers_by_block(layer_sensitivities)
        

        block_budgets = None
        if layer_activation_entropies:
            # 尝试进行block-level分配
            logger.info("  Level 1: Block-level budget allocation (based on activation entropy)")
            logger.info("  Level 2: Layer-level greedy allocation (within each block)")
            
            # 使用min_rank和max_rank作为block平均rank的范围
            block_budgets, block_avg_ranks = allocate_block_budgets_by_entropy(
                blocks, 
                layer_activation_entropies,
                min_avg_rank=min_rank,
                max_avg_rank=max_rank
            )
            logger.info("✓ Block-level budget allocation completed successfully")
            logger.info("  Now proceeding to layer-level allocation within each block...")

        else:
            logger.info("=" * 80)
            logger.info("SINGLE-LEVEL ALLOCATION (Block-level uniform)")
            logger.info(f"  Using uniform block budgets: num_layers × avg_rank={avg_rank}")
            logger.info("=" * 80)
        
        effective_layer_max_rank = layer_max_rank if layer_max_rank is not None else max_rank
        
        return allocate_ranks_block_greedy(
            blocks, entropy_reduction_tables, base_rank, step_size, effective_layer_max_rank, avg_rank,
            layer_activation_norms=layer_activation_norms,
            layer_activation_entropies=layer_activation_entropies,
            activation_entropy_beta=activation_entropy_beta,
            use_activation_entropy=use_activation_entropy,
            layer_stable_ranks=layer_stable_ranks,
            block_budgets=block_budgets 
        )

def get_model(model):
    import torch
    from transformers import AutoModelForCausalLM

    def skip(*args, **kwargs):
        pass

    torch.nn.init.kaiming_uniform_ = skip
    torch.nn.init.uniform_ = skip
    torch.nn.init.normal_ = skip
    model_lower = model.lower()
    model_type = ""
    try:
        cfg = AutoConfig.from_pretrained(model, trust_remote_code=True)
        model_type = (getattr(cfg, "model_type", "") or "").lower()
    except Exception as e:
        logger.warning(f"AutoConfig load failed for {model}, fallback to path-based type detection: {e}")

    if model_type in {"opt"} or ("opt" in model_lower and not model_type):
        from transformers import OPTForCausalLM
        model = OPTForCausalLM.from_pretrained(model, torch_dtype=torch.float16)
        model.seqlen = model.config.max_position_embeddings
    elif model_type.startswith("llama") or ("llama" in model_lower and not model_type):
        from transformers import LlamaForCausalLM
        model = LlamaForCausalLM.from_pretrained(model, torch_dtype=torch.float16)
        model.seqlen = model.config.max_position_embeddings
    elif model_type == "qwen3":
        try:
            from transformers import Qwen3ForCausalLM
            model = Qwen3ForCausalLM.from_pretrained(model, torch_dtype=torch.float16)
        except ImportError:
            raise RuntimeError(
                "Qwen3 model requires transformers>=4.51 in the active environment."
            )
        model.seqlen = model.config.max_position_embeddings
    elif (
        model_type.startswith("qwen2")
        or model_type == "qwen"
        or (not model_type and "qwen" in model_lower)
    ):
        # ⭐ 支持 Qwen2.x / Qwen2.5 模型
        try:
            from transformers import Qwen2ForCausalLM
            model = Qwen2ForCausalLM.from_pretrained(model, torch_dtype=torch.float16)
        except ImportError:
            model = AutoModelForCausalLM.from_pretrained(
                model, torch_dtype=torch.float16, trust_remote_code=True
            )
        model.seqlen = model.config.max_position_embeddings
    else:
        model = AutoModelForCausalLM.from_pretrained(
            model, torch_dtype=torch.float16, trust_remote_code=True
        )
        model.seqlen = model.config.max_position_embeddings
        logger.warning(
            f"Unsupported/unknown model_type={model_type}, loaded via AutoModelForCausalLM: {model}"
        )
    return model


CASCADE_ANALYSIS_TARGETS = [
    "q_proj",
    "k_proj",
    "v_proj",
    "o_proj",
    "gate_proj",
    "up_proj",
    "down_proj",
]
CASCADE_ANALYSIS_RANKS = [16, 32, 64, 128]
CASCADE_ANALYSIS_BITS = [2, 3, 4, 8]
CASCADE_ANALYSIS_STEPS = 4
CASCADE_ANALYSIS_OUTPUT = "qwen_svd_cascade_analysis.xlsx"
RESIDUAL_DIST_ANALYSIS_OUTPUT = "qwen_residual_distribution_analysis.xlsx"


def _cascade_fake_quantize_per_channel(tensor, bit):
    bit = int(bit)
    if bit == 2:
        quantized = quantize_to_2bit_perchannel(tensor.float())
        return quantized.float()
    if bit == 3:
        quantized, _, _ = quantize_to_3bit_perchannel(tensor.float())
        return quantized.float()
    if bit == 4:
        quantized, _, _ = quantize_to_4bit_perchannel(tensor.float())
        return quantized.float()
    if bit == 8:
        quantized, _, _ = quantize_to_8bit_perchannel(tensor.float())
        return quantized.float()
    raise ValueError(f"Unsupported analysis bit width: {bit}. Expected one of [2, 3, 4, 8].")


def _prepare_analysis_hessian(W, H):
    # Keep Hessian symmetric and numerically stable before quadratic-form error evaluation.
    H = 0.5 * (H.float() + H.float().t())
    diag = torch.diag(H)
    dead = diag <= 0
    if bool(dead.any()):
        H[dead, dead] = 1.0
        W = W.clone()
        W[:, dead] = 0
    return W, H


def _run_cascaded_svd_analysis(module, gptq_handler, layer_name):
    W = module.weight.data.clone()
    if isinstance(module, transformers.Conv1D):
        W = W.t()
    W = W.float()

    in_features = int(W.shape[1])
    out_features = int(W.shape[0])

    H_orig = gptq_handler.H.clone().float() if (hasattr(gptq_handler, "H") and gptq_handler.H is not None) else None
    if H_orig is None or H_orig.shape[0] != in_features or H_orig.shape[1] != in_features:
        logger.warning(
            f"[Cascade Analysis] layer={layer_name}: invalid Hessian shape "
            f"{None if H_orig is None else tuple(H_orig.shape)}, fallback to identity."
        )
        H_orig = torch.eye(in_features, device=W.device, dtype=W.dtype)

    W, H_orig = _prepare_analysis_hessian(W, H_orig)
    baseline_energy = torch.sum((W @ H_orig) * W).item()
    if baseline_energy <= 1e-12:
        baseline_energy = torch.sum(W * W).item() + 1e-12
        logger.warning(
            f"[Cascade Analysis] layer={layer_name}: baseline_energy is near zero under Hessian form, "
            "fallback to Frobenius baseline."
        )

    records = []
    with torch.inference_mode():
        for rank in CASCADE_ANALYSIS_RANKS:
            for bit in CASCADE_ANALYSIS_BITS:
                residual = W.clone()
                prev_norm_error = 1.0
                cumulative_cost = 0.0

                for step in range(1, CASCADE_ANALYSIS_STEPS + 1):
                    max_rank = min(int(residual.shape[0]), int(residual.shape[1]))
                    use_rank = min(int(rank), max_rank)
                    if use_rank <= 0:
                        break

                    U, S, Vh = torch.linalg.svd(residual, full_matrices=False)
                    low_rank = (U[:, :use_rank] * S[:use_rank]) @ Vh[:use_rank, :]
                    low_rank_q = _cascade_fake_quantize_per_channel(low_rank, bit)
                    residual = residual - low_rank_q

                    act_error = torch.sum((residual @ H_orig) * residual).item()
                    weight_error = torch.sum(residual * residual).item()
                    if act_error < 0 and abs(act_error) < 1e-9:
                        act_error = 0.0
                    norm_error = act_error / baseline_energy
                    delta_error = prev_norm_error - norm_error

                    # Cost is intentionally normalized to average bits per original weight.
                    step_cost = (
                        float(rank) * float(in_features + out_features) * float(bit)
                        / float(in_features * out_features)
                    )
                    cumulative_cost += step_cost
                    score = (delta_error / step_cost) if step_cost > 0 else 0.0

                    records.append(
                        {
                            "Layer_Name": layer_name,
                            "Rank": int(rank),
                            "Bit": int(bit),
                            "SVD_Step": int(step),
                            "Step_Bit_Cost": float(step_cost),
                            "Cumulative_Bit_Cost": float(cumulative_cost),
                            "Absolute_Weight_Error": float(weight_error),
                            "Absolute_Act_Error": float(act_error),
                            "Normalized_Act_Error": float(norm_error),
                            "Delta_Normalized_Error": float(delta_error),
                            "Marginal_Score": float(score),
                        }
                    )
                    prev_norm_error = norm_error

    return records


def _export_cascade_analysis_excel(layer_rows_map, output_path):
    import pandas as pd

    try:
        from openpyxl.styles import Font, PatternFill
        style_supported = True
    except Exception:
        style_supported = False

    wrote_sheet = False
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for layer_name in CASCADE_ANALYSIS_TARGETS:
            rows = layer_rows_map.get(layer_name, [])
            if not rows:
                continue

            df = pd.DataFrame(rows)
            df = df.sort_values(["Rank", "Bit", "SVD_Step"], kind="stable").reset_index(drop=True)
            best_per_step = df.groupby("SVD_Step")["Marginal_Score"].transform("max")
            df["Best_In_Step"] = (df["Marginal_Score"] >= (best_per_step - 1e-12))
            df.to_excel(writer, sheet_name=layer_name, index=False)
            wrote_sheet = True

            if style_supported:
                ws = writer.sheets[layer_name]
                headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
                best_col = headers.get("Best_In_Step")
                if best_col is not None:
                    bold_font = Font(bold=True)
                    highlight_fill = PatternFill(fill_type="solid", fgColor="FFFDEB84")
                    for r in range(2, ws.max_row + 1):
                        flag = ws.cell(row=r, column=best_col).value
                        if flag in (True, 1, "TRUE", "True", "true"):
                            for c in range(1, ws.max_column + 1):
                                cell = ws.cell(row=r, column=c)
                                cell.font = bold_font
                                cell.fill = highlight_fill
                ws.freeze_panes = "A2"

        if not wrote_sheet:
            empty_df = pd.DataFrame(
                columns=[
                    "Layer_Name",
                    "Rank",
                    "Bit",
                    "SVD_Step",
                    "Step_Bit_Cost",
                    "Cumulative_Bit_Cost",
                    "Absolute_Weight_Error",
                    "Absolute_Act_Error",
                    "Normalized_Act_Error",
                    "Delta_Normalized_Error",
                    "Marginal_Score",
                    "Best_In_Step",
                ]
            )
            empty_df.to_excel(writer, sheet_name="summary", index=False)

    logger.info(f"[Cascade Analysis] Excel exported to: {output_path}")


def _resolve_analysis_output_dir(path_like):
    out_dir = str(path_like).strip() if path_like is not None else ""
    if out_dir == "":
        out_dir = "BiLLM/residual_analysis"
    if not os.path.isabs(out_dir):
        out_dir = os.path.abspath(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


def _build_row_feature_dataframe(W, R):
    import pandas as pd

    eps = 1e-12

    w_mean = W.mean(dim=1)
    w_var = W.var(dim=1, unbiased=False)
    w_range = W.max(dim=1).values - W.min(dim=1).values
    w_median = W.median(dim=1).values
    w_abs_mean = W.abs().mean(dim=1)
    w_abs_median = W.abs().median(dim=1).values
    w_tc = w_abs_median / (w_abs_mean + eps)

    r_mean = R.mean(dim=1)
    r_var = R.var(dim=1, unbiased=False)
    r_range = R.max(dim=1).values - R.min(dim=1).values
    r_median = R.median(dim=1).values
    r_abs_mean = R.abs().mean(dim=1)
    r_abs_median = R.abs().median(dim=1).values
    r_tc = r_abs_median / (r_abs_mean + eps)

    n_rows = int(W.shape[0])
    df = pd.DataFrame({
        "Row_Index": list(range(n_rows)),
        "W_Row_Mean": w_mean.detach().cpu().numpy(),
        "W_Row_Var": w_var.detach().cpu().numpy(),
        "W_Row_Range": w_range.detach().cpu().numpy(),
        "W_Row_Median": w_median.detach().cpu().numpy(),
        "Tc_W_Row": w_tc.detach().cpu().numpy(),
        "R_Row_Mean": r_mean.detach().cpu().numpy(),
        "R_Row_Var": r_var.detach().cpu().numpy(),
        "R_Row_Range": r_range.detach().cpu().numpy(),
        "R_Row_Median": r_median.detach().cpu().numpy(),
        "Tc_R_Row": r_tc.detach().cpu().numpy(),
    })
    return df


def _plot_layer_3d_w_r(W, R, layer_name, save_dir):
    import numpy as np
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    w_np = W.detach().cpu().float().numpy()
    r_np = R.detach().cpu().float().numpy()

    rows, cols = w_np.shape
    max_points = 220
    step_r = max(1, math.ceil(rows / max_points))
    step_c = max(1, math.ceil(cols / max_points))

    w_s = w_np[::step_r, ::step_c]
    r_s = r_np[::step_r, ::step_c]

    y = np.arange(w_s.shape[0]) * step_r
    x = np.arange(w_s.shape[1]) * step_c
    X, Y = np.meshgrid(x, y)

    z_min = float(min(w_s.min(), r_s.min()))
    z_max = float(max(w_s.max(), r_s.max()))
    if abs(z_max - z_min) < 1e-12:
        z_max = z_min + 1e-6

    fig = plt.figure(figsize=(16, 6))
    ax1 = fig.add_subplot(1, 2, 1, projection='3d')
    ax2 = fig.add_subplot(1, 2, 2, projection='3d')

    ax1.plot_surface(X, Y, w_s, cmap='viridis', linewidth=0, antialiased=False)
    ax2.plot_surface(X, Y, r_s, cmap='plasma', linewidth=0, antialiased=False)

    for ax, title in ((ax1, 'W Distribution'), (ax2, 'Residual Distribution')):
        ax.set_xlabel('Column')
        ax.set_ylabel('Row')
        ax.set_zlabel('Value')
        ax.set_title(title)
        ax.set_zlim(z_min, z_max)

    fig.suptitle(f'Layer 1 - {layer_name} (shared z-scale)', fontsize=13, fontweight='bold')
    fig.tight_layout()

    out_path = os.path.join(save_dir, f'layer_1_{layer_name}_3d.png')
    fig.savefig(out_path, dpi=160, bbox_inches='tight')
    plt.close(fig)
    return out_path


def _plot_layer_density_w_r(W, R, layer_name, save_dir):
    import numpy as np
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    w_flat = W.detach().cpu().float().numpy().reshape(-1)
    r_flat = R.detach().cpu().float().numpy().reshape(-1)

    x_min = float(min(w_flat.min(), r_flat.min()))
    x_max = float(max(w_flat.max(), r_flat.max()))
    if abs(x_max - x_min) < 1e-12:
        x_max = x_min + 1e-6

    bins = 240
    w_hist, edges = np.histogram(w_flat, bins=bins, range=(x_min, x_max), density=True)
    r_hist, _ = np.histogram(r_flat, bins=bins, range=(x_min, x_max), density=True)
    centers = 0.5 * (edges[:-1] + edges[1:])

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(centers, w_hist, label='W Density', color='tab:blue', linewidth=1.5)
    ax.plot(centers, r_hist, label='Residual Density', color='tab:red', linewidth=1.5)
    ax.set_xlim(x_min, x_max)
    ax.set_xlabel('Value')
    ax.set_ylabel('Probability Density')
    ax.set_title(f'Layer 1 - {layer_name} (shared x-scale)')
    ax.grid(alpha=0.25)
    ax.legend()
    fig.tight_layout()

    out_path = os.path.join(save_dir, f'layer_1_{layer_name}_pdf.png')
    fig.savefig(out_path, dpi=180, bbox_inches='tight')
    plt.close(fig)
    return out_path


def _run_residual_distribution_analysis(module, gptq_handler, layer_name, args, layer_rank, output_dir):
    W = module.weight.data.clone()
    if isinstance(module, transformers.Conv1D):
        W = W.t()
    W = W.float()

    in_features = int(W.shape[1])
    H_orig = gptq_handler.H.clone().float() if (hasattr(gptq_handler, "H") and gptq_handler.H is not None) else None
    if H_orig is None or H_orig.shape[0] != in_features or H_orig.shape[1] != in_features:
        logger.warning(
            f"[Residual Dist] layer={layer_name}: invalid Hessian shape "
            f"{None if H_orig is None else tuple(H_orig.shape)}, fallback to identity."
        )
        H_orig = torch.eye(in_features, device=W.device, dtype=W.dtype)

    W_eval, H_eval = _prepare_analysis_hessian(W, H_orig)

    selected_rank = int(layer_rank)
    rank_trace = []

    with torch.inference_mode():
        if bool(getattr(args, "svd_enable", False)):
            if bool(getattr(args, "svd_auto_rank_deltae_stop", False)):
                selected_rank, rank_trace = gptq_handler._search_single_svd_rank_by_delta_norm_error(
                    W=W_eval,
                    H=H_eval,
                    svd_lowrank_fp16=bool(getattr(args, "svd_lowrank_fp16", False)),
                    start_rank=16,
                    rank_step=16,
                    delta_norm_threshold=0.05,
                )
                max_rank = min(int(W_eval.shape[0]), int(W_eval.shape[1]))
                selected_rank = int(max(1, min(int(selected_rank), max_rank)))
                _, _, proj_down, proj_up = gptq_handler._svd_decompose(W_eval, rank=selected_rank)
            else:
                max_rank = min(int(W_eval.shape[0]), int(W_eval.shape[1]))
                selected_rank = int(max(1, min(int(selected_rank), max_rank)))
                _, _, proj_down, proj_up = gptq_handler._svd_decompose(W_eval, rank=selected_rank)

            if bool(getattr(args, "svd_lowrank_fp16", False)):
                L_q = gptq_handler._compose_lowrank_fp16(proj_down, proj_up).float()
            else:
                L_q = gptq_handler._quantize_lowrank_4bit(proj_down, proj_up).float()
            R = (W_eval - L_q).float()
        else:
            logger.warning("[Residual Dist] svd_enable=False, fallback to Residual=W for plotting.")
            selected_rank = 0
            R = W_eval.clone()

    row_df = _build_row_feature_dataframe(W_eval, R)
    path_3d = _plot_layer_3d_w_r(W_eval, R, layer_name, output_dir)
    path_pdf = _plot_layer_density_w_r(W_eval, R, layer_name, output_dir)

    return {
        "row_stats": row_df,
        "selected_rank": int(selected_rank),
        "rank_trace": rank_trace,
        "plot_3d": path_3d,
        "plot_pdf": path_pdf,
    }


def _export_residual_distribution_excel(layer_results, output_path):
    import pandas as pd

    wrote_sheet = False
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for layer_name in CASCADE_ANALYSIS_TARGETS:
                info = layer_results.get(layer_name)
                if not info or "row_stats" not in info:
                    continue
                df = info["row_stats"]
                df.to_excel(writer, sheet_name=layer_name, index=False)
                wrote_sheet = True

            if not wrote_sheet:
                pd.DataFrame(columns=["Empty"]).to_excel(writer, sheet_name="summary", index=False)
    except Exception as exc:
        raise RuntimeError(
            f"Failed to export residual distribution Excel to {output_path}. "
            "Please ensure openpyxl is installed. "
            f"Original error: {exc}"
        )

    logger.info(f"[Residual Dist] Excel exported to: {output_path}")




'''
The function is employed to calibrate and quantize models layer by layer.
'''
@torch.no_grad()
def quant_sequential(model, dataloader, dev):
    logger.info("Starting ...")

    for name, module in model.named_modules():
        module.global_name = args.model + name

    use_cache = model.config.use_cache
    model.config.use_cache = False

    # Check model type using case-insensitive matching
    model_str = args.model.lower()
    layers = None
    
    if "opt" in model_str:
        layers = model.model.decoder.layers
        model.model.decoder.embed_tokens = model.model.decoder.embed_tokens.to(dev)
        model.model.decoder.embed_positions = model.model.decoder.embed_positions.to(dev)
        if hasattr(model.model.decoder, "project_out") and model.model.decoder.project_out:
            model.model.decoder.project_out = model.model.decoder.project_out.to(dev)
        if hasattr(model.model.decoder, "project_in") and model.model.decoder.project_in:
            model.model.decoder.project_in = model.model.decoder.project_in.to(dev)
    elif "llama" in model_str:
        layers = model.model.layers
        model.model.embed_tokens = model.model.embed_tokens.to(dev)
        model.model.norm = model.model.norm.to(dev)
        
        for layer in layers:
            if hasattr(layer, 'self_attn') and hasattr(layer.self_attn, 'rotary_emb'):
                layer.self_attn.rotary_emb = layer.self_attn.rotary_emb.to(dev)
                if hasattr(layer.self_attn.rotary_emb, 'inv_freq'):
                    layer.self_attn.rotary_emb.inv_freq = layer.self_attn.rotary_emb.inv_freq.to(dev)
    elif "qwen" in model_str:
        layers = model.model.layers
        model.model.embed_tokens = model.model.embed_tokens.to(dev)
        model.model.norm = model.model.norm.to(dev)
        
        for layer in layers:
            if hasattr(layer, 'self_attn') and hasattr(layer.self_attn, 'rotary_emb'):
                layer.self_attn.rotary_emb = layer.self_attn.rotary_emb.to(dev)
                if hasattr(layer.self_attn.rotary_emb, 'inv_freq'):
                    layer.self_attn.rotary_emb.inv_freq = layer.self_attn.rotary_emb.inv_freq.to(dev)
    else:
        raise ValueError(f"Unsupported model type: {args.model}.")
    
    layers[0] = layers[0].to(dev)

    dtype = next(iter(model.parameters())).dtype

    calibration_seqlen = min(model.seqlen, 2048)
    inps = torch.zeros(
        (args.nsamples, calibration_seqlen, model.config.hidden_size), dtype=dtype, device=dev
    )
    cache = {"i": 0, "attention_mask": None, "position_ids": None}

    class Catcher(nn.Module):
        def __init__(self, module):
            super().__init__()
            self.module = module
            # Qwen3 在外层循环会读取 decoder_layer.attention_type
            if hasattr(module, "attention_type"):
                self.attention_type = module.attention_type

        def __getattr__(self, name):
            try:
                return super().__getattr__(name)
            except AttributeError:
                wrapped = self._modules.get("module")
                if wrapped is not None and hasattr(wrapped, name):
                    return getattr(wrapped, name)
                raise

        def forward(self, inp, **kwargs):
            inps[cache["i"]] = inp
            cache["i"] += 1
            cache["attention_mask"] = kwargs.get("attention_mask", None)
            cache["position_ids"] = kwargs.get("position_ids", None)
            raise ValueError

    layers[0] = Catcher(layers[0])
    
    for batch in dataloader:
        try:
            inp = batch[0].to(dev)
            if inp.dim() == 1:
                inp = inp.unsqueeze(0)
            
            # 获取当前 batch 的 seq_len
            current_seq_len = inp.shape[1]
            
            # 创建 mask 和 position_ids
            attention_mask = torch.ones((inp.shape[0], current_seq_len), dtype=torch.long, device=dev)
            position_ids = torch.arange(0, current_seq_len, dtype=torch.long, device=dev)
            position_ids = position_ids.unsqueeze(0).expand(inp.shape[0], -1)

            # 调用模型
            model(input_ids=inp, attention_mask=attention_mask, position_ids=position_ids)
        except ValueError:
            pass
            
    layers[0] = layers[0].module
    layers[0] = layers[0].cpu()
    
    if "opt" in model_str:
        model.model.decoder.embed_tokens = model.model.decoder.embed_tokens.cpu()
        model.model.decoder.embed_positions = model.model.decoder.embed_positions.cpu()
        if hasattr(model.model.decoder, "project_out") and model.model.decoder.project_out:
            model.model.decoder.project_out = model.model.decoder.project_out.cpu()
        if hasattr(model.model.decoder, "project_in") and model.model.decoder.project_in:
            model.model.decoder.project_in = model.model.decoder.project_in.cpu()
    elif "llama" in model_str:
        model.model.embed_tokens = model.model.embed_tokens.cpu()
        model.model.norm = model.model.norm.cpu()
    elif "qwen" in model_str:
        # ⭐ 支持 Qwen2.5 模型（结构与 Llama 相同）
        model.model.embed_tokens = model.model.embed_tokens.cpu()
        model.model.norm = model.model.norm.cpu()
    
    torch.cuda.empty_cache()

    outs = torch.zeros_like(inps)
    attention_mask = cache["attention_mask"]
    position_ids = cache.get("position_ids", None)
    model_type = model.config.model_type.lower() if hasattr(model.config, "model_type") else ""
    use_shared_position_embeddings = (
        hasattr(model, "model")
        and hasattr(model.model, "rotary_emb")
        and (("llama" in model_type) or ("qwen" in model_type))
    )
    
    def forward_layer_with_position(layer, inp, attention_mask, position_ids=None):
        """向后兼容的 layer forward 调用"""
        pos_ids = position_ids
        if pos_ids is None:
            batch_size, seq_len = inp.shape[0], inp.shape[1]
            pos_ids = torch.arange(seq_len, dtype=torch.long, device=inp.device)
            pos_ids = pos_ids.unsqueeze(0).expand(batch_size, -1)
        else:
            pos_ids = pos_ids.to(inp.device)

        if use_shared_position_embeddings:
            try:
                # transformers>=4.55 的 Llama/Qwen layer forward 期望共享的 (cos, sin)
                position_embeddings = model.model.rotary_emb(inp, pos_ids)
                return layer(
                    inp,
                    attention_mask=attention_mask,
                    position_ids=pos_ids,
                    position_embeddings=position_embeddings,
                )[0]
            except TypeError:
                pass
        try:
            # 尝试使用 position_ids（Llama 3.1）
            return layer(inp, attention_mask=attention_mask, position_ids=pos_ids)[0]
        except TypeError:
            # 回退到不使用 position_ids（Llama 2）
            return layer(inp, attention_mask=attention_mask)[0]

    logger.info("Ready.")
    
    rank_allocations = {}
    svd_energy_events = []
    svd_bit_stats_events = []
    layer_avg_weight_bits_events = []
    svd_obr_scheme = str(getattr(args, "svd_obr_scheme", "none")).strip().lower()
    cascade_analysis_enabled = bool(getattr(args, "svd_cascade_analysis", False))
    cascade_analysis_rows = {name: [] for name in CASCADE_ANALYSIS_TARGETS}
    if cascade_analysis_enabled:
        logger.info(
            "[Cascade Analysis] Enabled. Hardcoded config: layer=0, "
            f"targets={CASCADE_ANALYSIS_TARGETS}, ranks={CASCADE_ANALYSIS_RANKS}, "
            f"bits={CASCADE_ANALYSIS_BITS}, steps={CASCADE_ANALYSIS_STEPS}."
        )

    residual_dist_analysis_enabled = bool(getattr(args, "svd_residual_distribution_analysis", False))
    residual_dist_layer_results = {name: None for name in CASCADE_ANALYSIS_TARGETS}
    residual_dist_output_dir = None
    if residual_dist_analysis_enabled:
        residual_dist_save_dir = getattr(args, "analyze_save_dir", "residual_analysis")
        if str(residual_dist_save_dir).strip() == "residual_analysis":
            residual_dist_save_dir = "BiLLM/residual_analysis"
        residual_dist_output_dir = _resolve_analysis_output_dir(residual_dist_save_dir)
        logger.info(
            "[Residual Dist] Enabled. Hardcoded config: layer=0, "
            f"targets={CASCADE_ANALYSIS_TARGETS}, output_dir={residual_dist_output_dir}."
        )

    if getattr(args, 'adaptive_rank', False) and getattr(args, 'svd_enable', False):
        logger.info("Adaptive Rank: Starting profiling phase...")
        
        layer_sensitivities = {}
        total_layers = 0
        entropy_reduction_tables = {}  # 存储熵减表（保留用于日志对比）
        energy_reduction_tables = {}  # 存储能量表（MSE表，用于绝对能量增益）
        layer_activation_norms = {}  # 存储激活范数（用于激活加权的能量增益分配）
        layer_activation_entropies = {}  # 存储激活熵（用于基于熵减的分配）
        layer_stable_ranks = {}  # 存储稳定秩（Stable Rank，用于衡量矩阵复杂度）
        layer_param_counts = {}  # 存储参数数量（保留用于日志，不用于计算）
        
        block_greedy_rank = getattr(args, 'block_greedy_rank', False)
        
        # 第一遍：收集层列表和激活信息
        # block_greedy_rank 模式：收集层列表和激活范数（用于激活加权的熵减分配）
        logger.info("Block Greedy Rank mode: Collecting layer list and activation norms...")
        for i in range(len(layers)):
            layer = layers[i].to(dev)
            subset = find_layers(layer)
            
            gptq_probe = {}
            for name in subset:
                if (
                    not (args.minlayer <= i < args.maxlayer and args.quant_only in name)
                ) == (not args.invert):
                    continue
                
                # 只处理 Linear 层
                is_linear = isinstance(subset[name], nn.Linear)
                if not is_linear:
                    try:
                        transformers_module = globals().get('transformers')
                        if transformers_module is not None:
                            is_linear = isinstance(subset[name], transformers_module.Conv1D)
                    except (AttributeError, TypeError):
                        pass
                if not is_linear:
                    continue
                
                braq_quantizer = Binarization(
                    subset[name].weight,
                    method=args.low_quant_method,
                    groupsize=groupsize,
                    iterative_2bit_enable=(
                        args.low_quant_method in {"2bit", "3bit", "4bit"}
                        and not getattr(args, "disable_iterative_2bit", False)
                    ),
                    iterative_2bit_iters=getattr(args, "iterative_2bit_iters", 5),
                    iterative_2bit_eps=getattr(args, "iterative_2bit_eps", 1e-8),
                    iterative_2bit_update_offset=(
                        not getattr(args, "iterative_2bit_disable_offset_update", False)
                    ),
                    iterative_2bit_fixed_iters=getattr(args, "iterative_2bit_fixed_iters", False),
                    iterative_2bit_print_mse=getattr(args, "iterative_2bit_print_mse", False),
                )
                gptq_probe[name] = BRAGPTQ(
                    subset[name],
                    braq_quantizer,
                    salient_metric=args.salient_metric,
                    disable_gptq=args.disable_gptq,
                )
                gptq_probe[name].profiling_mode = True
            
            if not gptq_probe:
                layer = layer.cpu()
                continue
            
            # ⭐ 收集完整激活值用于计算激活熵
            activation_storage = {}  # {name: list of activation tensors}
            for name in gptq_probe:
                activation_storage[name] = []
            
            def add_batch_probe(name):
                def tmp(_, inp, out):
                    gptq_probe[name].add_batch(inp[0].data, out.data)
                    # ⭐ 收集完整激活值用于计算激活熵
                    activation_storage[name].append(inp[0].data.clone())
                return tmp
            
            handles = []
            for name in gptq_probe:
                handles.append(subset[name].register_forward_hook(add_batch_probe(name)))
            
            for j in range(args.nsamples):
                outs[j] = forward_layer_with_position(layer, inps[j].unsqueeze(0), attention_mask, position_ids)
            
            for h in handles:
                h.remove()
            
            for name in gptq_probe:
                # 计算平均激活范数
                if len(gptq_probe[name].activation_norms) > 0:
                    avg_activation_norm = sum(gptq_probe[name].activation_norms) / len(gptq_probe[name].activation_norms)
                else:
                    avg_activation_norm = 1.0
  
                activation_entropy = 1.0  # 默认值
                if len(activation_storage[name]) > 0:
                    try:
                        all_activations = torch.cat(activation_storage[name], dim=0)  # shape: [N, d]
                        feature_stds = torch.std(all_activations, dim=0)  # shape: [d]

                        epsilon = 1e-8
                        d = feature_stds.shape[0]
                        if d > 0:
                            log_stds = torch.log(feature_stds + epsilon)
                            mean_log_std = torch.mean(log_stds)  # (1/d) * Σ_j log(σ_j + ε)
                            activation_entropy = torch.exp(mean_log_std).item()  # exp(...) 确保为正数
                    except Exception as e:
                        logger.warning(f"Layer {i} {name}: failed to compute activation_entropy: {e}, using default 1.0")
                        activation_entropy = 1.0
                else:
                    logger.warning(f"Layer {i} {name}: no activation values collected, using default activation_entropy=1.0")
                
                
                layer_key = f"{i}.{name}"
                layer_sensitivities[layer_key] = 1.0  # 占位符，实际不使用
                layer_activation_norms[layer_key] = avg_activation_norm
                layer_activation_entropies[layer_key] = activation_entropy
                total_layers += 1
                
                # ⭐ 打印激活和参数收集信息（INFO级别，便于调试）
                logger.info(f"Layer {i} {name}: collected activation_norm={avg_activation_norm:.6f}, "
                            f"activation_entropy={activation_entropy:.6f} "
                            f"(from {len(gptq_probe[name].activation_norms)} samples), "
                            )
            
            layer = layer.cpu()
            del gptq_probe, handles
            torch.cuda.empty_cache()

        # 如果启用 block_greedy_rank，构建熵减表
        if block_greedy_rank:
            logger.info("Block Greedy Rank: Building entropy reduction tables...")
            
            from tqdm import tqdm

            
            base_rank = getattr(args, 'block_greedy_base_rank', 128)
            max_rank_svd = getattr(args, 'block_greedy_max_rank', 512)
            step_size = getattr(args, 'block_greedy_step_size', 32)
            
            
            logger.info(f"Building entropy tables: base_rank={base_rank}, "
                       f"max_rank_svd={max_rank_svd}, step_size={step_size}")
            
            for i in tqdm(range(len(layers)), desc="Building entropy tables"):
                layer = layers[i].to(dev)
                subset = find_layers(layer)
                
                for name in subset:
                    if (
                        not (args.minlayer <= i < args.maxlayer and args.quant_only in name)
                    ) == (not args.invert):
                        continue
                    
                    # 只处理 Linear 层
                    is_linear = isinstance(subset[name], nn.Linear)
                    if not is_linear:
                        try:
                            transformers_module = globals().get('transformers')
                            if transformers_module is not None:
                                is_linear = isinstance(subset[name], transformers_module.Conv1D)
                        except (AttributeError, TypeError):
                            pass
                    if not is_linear:
                        continue
                    
                    layer_key = f"{i}.{name}"
                    if layer_key not in layer_sensitivities:
                        continue
                    
                    # 获取权重矩阵
                    W = subset[name].weight.data.clone()
                    if isinstance(subset[name], transformers.Conv1D):
                        W = W.t()
                    W = W.float()
                    
                    import torch.linalg as LA
                    S = LA.svdvals(W)

                    
                    # ⭐ 计算稳定秩（Stable Rank）：r_stable = ||W||_F^2 / ||W||_2^2 = Σ σ_i^2 / σ_max^2
                    # 稳定秩衡量矩阵复杂度：q/k 的稳定秩低（信息简单，谱分布陡峭），v/mlp 的稳定秩高（信息复杂，谱分布平缓）
                    S_squared_for_stable = S ** 2
                    frobenius_norm_squared = torch.sum(S_squared_for_stable).item()  # ||W||_F^2 = Σ σ_i^2
                    spectral_norm_squared = S[0].item() ** 2 if len(S) > 0 else 1.0  # ||W||_2^2 = σ_max^2
                    if spectral_norm_squared > 1e-9:
                        stable_rank = frobenius_norm_squared / spectral_norm_squared
                    else:
                        stable_rank = 1.0
                        logger.warning(f"Layer {layer_key}: spectral_norm_squared too small ({spectral_norm_squared:.6e}), using default stable_rank=1.0")
                    
                    layer_stable_ranks[layer_key] = stable_rank
                    
                    entropy_table = {}
                    num_singular_values = len(S)
                    
                    
                    max_rank_actual = min(max_rank_svd, num_singular_values)
                    S_squared = S ** 2
                    

                    mse_base = torch.sum(S_squared[base_rank:]).item()
                    
                    entropy_table[base_rank] = 0.0  # 基准点
                    
                    # 计算各个 rank 的累积熵减（只涉及向量切片和求和，零开销）
                    for rank in range(base_rank + step_size, max_rank_actual + 1, step_size):
                        # rank 对应的 MSE = 从 rank 开始到最后的奇异值平方和

                        mse_r = torch.sum(S_squared[rank:]).item()
                        
                        logger.info(f"Layer {layer_key}: mse_base={mse_base}, mse_r={mse_r}")
                        entropy_reduction = 0.5 * torch.log(torch.tensor(mse_base / mse_r)).item()
                        entropy_table[rank] = entropy_reduction
                    
                    if max_rank_actual not in entropy_table:
                        if max_rank_actual >= num_singular_values:
                            mse_max = 0.0
                        else:
                            mse_max = torch.sum(S_squared[max_rank_actual:]).item()
                        
                        
                        logger.info(f"Layer {layer_key}: mse_base={mse_base}, mse_max={mse_max}")
                        entropy_reduction = 0.5 * torch.log(torch.tensor(mse_base / mse_max)).item() 
                        entropy_table[max_rank_actual] = entropy_reduction
                    
                    entropy_reduction_tables[layer_key] = entropy_table 
                    
                    logger.debug(f"Layer {layer_key}: entropy_table built with {len(entropy_table)} entries, "
                               f"stable_rank={stable_rank:.6f}")
                    
                    # 清理中间变量（立即清理，节省内存）
                    del S, S_squared, W
                    torch.cuda.empty_cache()
                
                layer = layer.cpu()
                torch.cuda.empty_cache()
            

            missing_entropy_tables = []
            for layer_key in layer_sensitivities.keys():
                if layer_key not in entropy_reduction_tables:
                    missing_entropy_tables.append(layer_key)
            

            logger.info(f"Entropy reduction tables built for {len(entropy_reduction_tables)} layers (for comparison)")
            logger.info(f"Stable ranks computed for {len(layer_stable_ranks)} layers")
        
        if layer_sensitivities:
            total_budget = total_layers * getattr(args, 'avg_rank', 256)
            min_rank = getattr(args, 'min_rank', 128)
            max_rank = getattr(args, 'max_rank', 512)
            
            logger.info(f"Total budget: {total_budget} (layers={total_layers}, avg_rank={getattr(args, 'avg_rank', 256)})")
            logger.info(f"Rank range: [{min_rank}, {max_rank}]")
            
            # 根据是否启用 block_greedy_rank 选择分配策略
            block_greedy_rank = getattr(args, 'block_greedy_rank', False)
            if block_greedy_rank:
                base_rank = getattr(args, 'block_greedy_base_rank', 128)
                step_size = getattr(args, 'block_greedy_step_size', 32)
                avg_rank = getattr(args, 'avg_rank', 256)
                # 获取激活熵相关参数
                activation_entropy_beta = getattr(args, 'activation_entropy_beta', 2.0)
                use_activation_entropy = getattr(args, 'use_activation_entropy', False)
                
                layer_max_rank = getattr(args, 'block_greedy_max_rank', 512)
                
                # ⭐ 控制是否使用两阶段分配策略
                disable_two_level = getattr(args, 'disable_two_level_allocation', False)
                if disable_two_level:
                    # 禁用两阶段分配：将 layer_activation_entropies 设为 None
                    # 这样会使用单阶段分配（每个块使用统一的预算：num_layers × avg_rank）
                    layer_activation_entropies_for_allocation = None
                    logger.info("Two-level allocation disabled: using single-level allocation (uniform block budgets)")
                    logger.info(f"  Layer rank range: [{base_rank}, {layer_max_rank}]")
                else:
                    # 启用两阶段分配：传入 layer_activation_entropies 用于块级别预算分配
                    layer_activation_entropies_for_allocation = layer_activation_entropies
                    logger.info(f"  Level 1 (Block): avg_rank range: [{min_rank}, {max_rank}]")
                    logger.info(f"  Level 2 (Layer): rank range: [{base_rank}, {layer_max_rank}]")
                
                rank_allocations = allocate_ranks(
                    layer_sensitivities, total_budget, min_rank, max_rank,
                    block_greedy=True,
                    entropy_reduction_tables=entropy_reduction_tables,
                    base_rank=base_rank,
                    step_size=step_size,
                    avg_rank=avg_rank,
                    layer_activation_norms=layer_activation_norms,
                    layer_activation_entropies=layer_activation_entropies_for_allocation,
                    activation_entropy_beta=activation_entropy_beta,
                    use_activation_entropy=use_activation_entropy,
                    layer_stable_ranks=layer_stable_ranks,
                    layer_max_rank=layer_max_rank  
                )
            else:
                rank_allocations = allocate_ranks(layer_sensitivities, total_budget, min_rank, max_rank)
            
            # 输出分配结果
            logger.info("Rank allocations:")
            for layer_key, allocated_rank in sorted(rank_allocations.items()):
                if block_greedy_rank:
                    logger.info(f"  {layer_key}: rank={allocated_rank}")
                else:
                    sensitivity = layer_sensitivities[layer_key]
                    logger.info(f"  {layer_key}: rank={allocated_rank}, sensitivity={sensitivity:.6f}")
        
        logger.info("Adaptive Rank: Profiling phase completed.")
    
    # === 阶段 2: 正式量化阶段 ===
    for i in range(len(layers)):
        layer = layers[i].to(dev)

        subset = find_layers(layer)
        selected_names = []
        for name in subset:
            if (
                not (args.minlayer <= i < args.maxlayer and args.quant_only in name)
            ) == (not args.invert):
                continue
            selected_names.append(name)

        use_true_sequential = bool(
            getattr(args, "gptq_true_sequential", False)
            and getattr(args, "std_gptq_enable", False)
        )
        if use_true_sequential:
            sequential_template = [
                ["self_attn.k_proj", "self_attn.v_proj", "self_attn.q_proj"],
                ["self_attn.o_proj"],
                ["mlp.up_proj", "mlp.gate_proj"],
                ["mlp.down_proj"],
            ]
            remaining = set(selected_names)
            sequential_groups = []
            for group in sequential_template:
                cur = [n for n in group if n in remaining]
                if cur:
                    sequential_groups.append(cur)
                    for n in cur:
                        remaining.remove(n)
            for n in selected_names:
                if n in remaining:
                    sequential_groups.append([n])
                    remaining.remove(n)
            if getattr(args, "std_gptq_enable", False):
                mode_name = "std_gptq"
            else:
                mode_name = "GPTQ"
            logger.info(
                f"Layer {i}: {mode_name} true-sequential groups = "
                + " | ".join([",".join(g) for g in sequential_groups])
            )
        else:
            sequential_groups = [selected_names]

        for group_names in sequential_groups:
            gptq = {}
            for name in group_names:
                braq_quantizer = Binarization(
                    subset[name].weight,
                    method=args.low_quant_method,
                    groupsize=groupsize,
                    iterative_2bit_enable=(
                        args.low_quant_method in {"2bit", "3bit", "4bit"}
                        and not getattr(args, "disable_iterative_2bit", False)
                    ),
                    iterative_2bit_iters=getattr(args, "iterative_2bit_iters", 5),
                    iterative_2bit_eps=getattr(args, "iterative_2bit_eps", 1e-8),
                    iterative_2bit_update_offset=(
                        not getattr(args, "iterative_2bit_disable_offset_update", False)
                    ),
                    iterative_2bit_fixed_iters=getattr(args, "iterative_2bit_fixed_iters", False),
                    iterative_2bit_print_mse=getattr(args, "iterative_2bit_print_mse", False),
                )
                gptq[name] = BRAGPTQ(
                    subset[name],
                    braq_quantizer,
                    salient_metric=args.salient_metric,
                    disable_gptq=args.disable_gptq,
                )
                # 确保正式量化阶段不是探针模式
                gptq[name].profiling_mode = False

            if not gptq:
                continue

            def add_batch(name):
                def tmp(_, inp, out):
                    gptq[name].add_batch(inp[0].data, out.data)
                return tmp

            handles = []
            for name in gptq:
                handles.append(subset[name].register_forward_hook(add_batch(name)))

            for j in range(args.nsamples):
                outs[j] = forward_layer_with_position(layer, inps[j].unsqueeze(0), attention_mask, position_ids)

            for h in handles:
                h.remove()

            for name in gptq:
                logger.info(f"{i} {name}")
                logger.info("Quantizing ...")
                
                # 确定使用的 rank（自适应或固定）
                layer_key = f"{i}.{name}"
                if getattr(args, 'adaptive_rank', False) and layer_key in rank_allocations:
                    layer_rank = rank_allocations[layer_key]
                    logger.info(f"Using adaptive rank: {layer_rank} (allocated for layer {layer_key})")
                else:
                    layer_rank = args.svd_rank

                if cascade_analysis_enabled and i == 0:
                    short_name = name.split(".")[-1]
                    if short_name in cascade_analysis_rows:
                        rows = _run_cascaded_svd_analysis(
                            module=subset[name],
                            gptq_handler=gptq[name],
                            layer_name=short_name,
                        )
                        cascade_analysis_rows[short_name].extend(rows)
                        logger.info(
                            f"[Cascade Analysis] layer={short_name}: collected {len(rows)} rows "
                            f"(expected {len(CASCADE_ANALYSIS_RANKS) * len(CASCADE_ANALYSIS_BITS) * CASCADE_ANALYSIS_STEPS})."
                        )
                    else:
                        logger.info(f"[Cascade Analysis] skip non-target module: {name}")
                    gptq[name].free()
                    continue

                if residual_dist_analysis_enabled and i == 0:
                    short_name = name.split(".")[-1]
                    if short_name in residual_dist_layer_results:
                        info = _run_residual_distribution_analysis(
                            module=subset[name],
                            gptq_handler=gptq[name],
                            layer_name=short_name,
                            args=args,
                            layer_rank=layer_rank,
                            output_dir=residual_dist_output_dir,
                        )
                        residual_dist_layer_results[short_name] = info
                        logger.info(
                            f"[Residual Dist] layer={short_name}: selected_rank={int(info.get('selected_rank', 0))}, "
                            f"rows={len(info.get('row_stats', []))}, "
                            f"plot_3d={info.get('plot_3d')}, plot_pdf={info.get('plot_pdf')}"
                        )
                    else:
                        logger.info(f"[Residual Dist] skip non-target module: {name}")
                    gptq[name].free()
                    continue
                
                if args.analyze_residual and i in args.analyze_layers:
                    from residual_analyzer import save_residual
                    # import transformers
                    W = subset[name].weight.data.clone()
                    if isinstance(subset[name], transformers.Conv1D):
                        W = W.t()
                    W = W.float()
                    
                    # 在fasterquant之前保存H（因为fasterquant会删除H）
                    H = gptq[name].H.clone() if hasattr(gptq[name], 'H') and gptq[name].H is not None else None
                    if H is None:
                        # 如果H不存在，创建一个单位矩阵（简化处理）
                        print(f"H does not exist for layer {i} {name}, creating a unit matrix.")
                        H = torch.eye(W.shape[1], device=W.device, dtype=W.dtype)
                    
                    dead = torch.diag(H) == 0
                    H[dead, dead] = 1
                    
                    # 计算R（在量化之前）
                    if args.svd_enable:
                        L, _, proj_down, proj_up = gptq[name]._svd_decompose(W, rank=layer_rank)
                        L_quantized = gptq[name]._quantize_lowrank_4bit(proj_down, proj_up)
                        R = (W - L_quantized.float()).float()
                    else:
                        R = W.clone()
                    
                    R[:, dead] = 0
                    save_residual(i, name, R, H, args.svd_enable)
                
                # Handle svd_early_stop: default True, but can be disabled with --no_svd_early_stop
                svd_early_stop = getattr(args, 'svd_early_stop', True)
                if getattr(args, 'no_svd_early_stop', False):
                    svd_early_stop = False
                
                fq_kwargs = dict(
                    percdamp=args.percdamp,
                    blocksize=args.blocksize,
                    svd_rank=layer_rank,
                    svd_enable=args.svd_enable,
                    svd_auto_rank_deltae_stop=getattr(args, 'svd_auto_rank_deltae_stop', False),
                    svd_rank_by_vh_range_sigma=False,
                    svd_rank_by_vh_range_sigma_topk=32,
                    svd_lowrank_fp16=getattr(args, 'svd_lowrank_fp16', False),
                    svd_num_iters=getattr(args, 'svd_num_iters', 1),
                    svd_2bit_stages=getattr(args, 'svd_2bit_stages', 2),
                    svd_2bit_extra_8bit_stage_enable=getattr(args, 'svd_2bit_extra_8bit_stage_enable', False),
                    svd_2bit_extra_8bit_stage_rank=getattr(args, 'svd_2bit_extra_8bit_stage_rank', 64),
                    svd_binary_refit_enable=getattr(args, 'svd_binary_refit_enable', False),
                    svd_binary_refit_epochs=getattr(args, 'svd_binary_refit_epochs', 5),
                    svd_binary_refit_fp_epochs=getattr(args, 'svd_binary_refit_fp_epochs', 3),
                    svd_binary_refit_lr=getattr(args, 'svd_binary_refit_lr', 1e-3),
                    svd_binary_refit_min_lr=getattr(args, 'svd_binary_refit_min_lr', 1e-4),
                    svd_binary_refit_weight_decay=getattr(args, 'svd_binary_refit_weight_decay', 1e-5),
                    svd_binary_refit_grad_clip=getattr(args, 'svd_binary_refit_grad_clip', 1.0),
                    svd_binary_refit_patience=getattr(args, 'svd_binary_refit_patience', 2),
                    svd_binary_refit_max_tokens=getattr(args, 'svd_binary_refit_max_tokens', 128),
                    svd_2bit_salient_3bit_enable=getattr(args, 'svd_2bit_salient_3bit_enable', False),
                    svd_2bit_salient_4bit_enable=getattr(args, 'svd_2bit_salient_4bit_enable', False),
                    svd_2bit_obr_cascade_enable=getattr(args, 'svd_2bit_obr_cascade_enable', False),
                    svd_2bit_obr_cascade_4bit_ratio=getattr(args, 'svd_2bit_obr_cascade_4bit_ratio', 0.1),
                    svd_2bit_obr_cascade_3bit_ratio=getattr(args, 'svd_2bit_obr_cascade_3bit_ratio', 0.1),
                    svd_2bit_obr_twogroup_enable=getattr(args, 'svd_2bit_obr_twogroup_enable', False),
                    svd_2bit_obr_twogroup_4bit_ratio=getattr(args, 'svd_2bit_obr_twogroup_4bit_ratio', 0.1),
                    svd_2bit_obr_twogroup_adaptive_enable=getattr(args, 'svd_2bit_obr_twogroup_adaptive_enable', False),
                    svd_2bit_obr_twogroup_adaptive_max_stages=getattr(args, 'svd_2bit_obr_twogroup_adaptive_max_stages', 3),
                    svd_2bit_obr_twogroup_adaptive_e128_threshold=getattr(args, 'svd_2bit_obr_twogroup_adaptive_e128_threshold', 0.2),
                    svd_2bit_obr_twogroup_adaptive_base_ratio=getattr(args, 'svd_2bit_obr_twogroup_adaptive_base_ratio', 0.1),
                    svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio=getattr(args, 'svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio', False),
                    svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor=getattr(args, 'svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor', False),
                    svd_2bit_obr_4group_enable=getattr(args, 'svd_2bit_obr_4group_enable', False),
                    svd_2bit_obr_4group_4bit_ratio=getattr(args, 'svd_2bit_obr_4group_4bit_ratio', 0.1),
                    svd_2bit_obr_4group_2bit_ratio_1=getattr(args, 'svd_2bit_obr_4group_2bit_ratio_1', 0.3),
                    svd_2bit_obr_4group_2bit_ratio_2=getattr(args, 'svd_2bit_obr_4group_2bit_ratio_2', 0.3),
                    svd_2bit_obr_4group_non4bit_scheme=getattr(args, 'svd_2bit_obr_4group_non4bit_scheme', 'order2'),
                    svd_2bit_obr_ternary_delta_factor=getattr(args, 'svd_2bit_obr_ternary_delta_factor', 0.7),
                    svd_obr_interblock_gptq_propagation_disable=getattr(args, 'svd_obr_interblock_gptq_propagation_disable', False),
                    svd_obr_cascade_lowrank_8bit_enable=getattr(args, 'svd_obr_cascade_lowrank_8bit_enable', False),
                    svd_obr_cascade_lowrank_gamma=getattr(args, 'svd_obr_cascade_lowrank_gamma', 1.0),
                    svd_obr_cascade_lowrank_damp=getattr(args, 'svd_obr_cascade_lowrank_damp', 0.01),
                    svd_2bit_salient_3bit_up_lim=getattr(args, 'svd_2bit_salient_3bit_up_lim', 50),
                    svd_2bit_salient_3bit_ratio=getattr(args, 'svd_2bit_salient_3bit_ratio', 0.1),
                    svd_2bit_salient_4bit_alpha=getattr(args, 'svd_2bit_salient_4bit_alpha', 0.4),
                    svd_2bit_salient_4bit_adaptive=getattr(args, 'svd_2bit_salient_4bit_adaptive', False),
                    svd_early_stop=svd_early_stop,
                    use_bass=False,
                    bass_pool_factor=None,
                    bass_batch_size=None,
                    bass_selection_mode="greedy",
                    bass_proxy_metric="linf",
                    disable_salient_mask=getattr(args, 'disable_salient_mask', False),
                    row_wise_split=getattr(args, 'row_wise_split', False),
                    column=getattr(args, 'column', False),
                    quant_method=args.low_quant_method,
                    obr=getattr(args, 'obr', False),
                    binary_residual=getattr(args, 'binary_residual', False),
                    svd_drop_residual=getattr(args, 'svd_drop_residual', False),
                    smoothquant_enable=getattr(args, 'smoothquant_enable', False),
                    smoothquant_alpha=getattr(args, 'smoothquant_alpha', 0.5),
                    smoothquant_beta=getattr(args, 'smoothquant_beta', -1),
                    smoothquant_span_mode=getattr(args, 'smoothquant_span_mode', 'absmax'),
                    salient_first_enable=getattr(args, 'salient_first_enable', False),
                    salient_first_ratio=getattr(args, 'salient_first_ratio', 0.1),
                    non_salient_svd_rank=getattr(args, 'non_salient_svd_rank', 512),
                    bias_correction=getattr(args, 'bias_correction', False),
                    linf_sigma_weight=1.0,
                    linf_vh_weight=1.0,
                    svd_one_opt=getattr(args, 'svd_one_opt', False),
                    svd_r_first=getattr(args, 'svd_r_first', False),
                    svd_r_first_refit_l=getattr(args, 'svd_r_first_refit_l', False),
                    structure_prune=getattr(args, 'structure_prune', False),
                    prune_n=getattr(args, 'prune_n', 4),
                    prune_m=getattr(args, 'prune_m', 8),
                    rotation=getattr(args, 'rotation', False),
                    svd_energy_stats_enable=getattr(args, 'svd_energy_stats_enable', False),
                    svd_energy_probe_rank=getattr(args, 'svd_energy_probe_rank', 128),
                    svd_energy_thresholds=getattr(args, 'svd_energy_thresholds', [0.9, 0.92]),
                    svd_bit_stats_enable=getattr(args, 'svd_bit_stats_enable', False),
                    svd_bit_stats_scale_bits=getattr(args, 'svd_bit_stats_scale_bits', 16),
                    svd_bit_stats_zp_bits=getattr(args, 'svd_bit_stats_zp_bits', 8),
                    svd_bit_stats_binary_data_bits=getattr(args, 'svd_bit_stats_binary_data_bits', 2.0),
                    svd_lwc_enable=getattr(args, 'svd_lwc_enable', False),
                    svd_lwc_candidates=getattr(args, 'svd_lwc_candidates', [1.0, 0.999, 0.995, 0.99, 0.98]),
                    svd_row_prune_search_enable=getattr(args, 'svd_row_prune_search_enable', False),
                    svd_row_prune_m_candidates=getattr(args, 'svd_row_prune_m_candidates', [0, 4, 8, 16, 32, 64]),
                    svd_row_prune_score_metric=getattr(args, 'svd_row_prune_score_metric', 'magnitude'),
                    svd_row_prune_binary_order=getattr(args, 'svd_row_prune_binary_order', 2),
                    svd_row_prune_quant_scheme=getattr(args, 'svd_row_prune_quant_scheme', 'binary'),
                    svd_row_prune_clip_search_enable=getattr(args, 'svd_row_prune_clip_search_enable', False),
                    svd_row_prune_clip_candidates=getattr(args, 'svd_row_prune_clip_candidates', [1.0, 0.999, 0.995, 0.99, 0.98, 0.96]),
                    svd_row_prune_clip_min_value=getattr(args, 'svd_row_prune_clip_min_value', 1e-8),
                    svd_row_prune_act_topk=getattr(args, 'svd_row_prune_act_topk', 3),
                    svd_row_prune_offline_refill_enable=getattr(args, 'svd_row_prune_offline_refill_enable', False),
                    svd_row_prune_offline_refill_sweeps=getattr(args, 'svd_row_prune_offline_refill_sweeps', 1),
                    svd_row_prune_offline_refill_max_positions=getattr(args, 'svd_row_prune_offline_refill_max_positions', -1),
                    svd_row_prune_offline_refill_include_zero=getattr(args, 'svd_row_prune_offline_refill_include_zero', True),
                    pure_gptq=False,
                    svd_then_pure_gptq=(
                        bool(getattr(args, 'svd256_then_pure_gptq2bit', False))
                        or bool(getattr(args, 'svd256_then_pure_gptq3bit', False))
                    ),
                    gptq_groupsize=getattr(args, 'gptq_groupsize', -1),
                    gptq_static_groups=getattr(args, 'gptq_static_groups', False),
                    gptq_act_order=getattr(args, 'gptq_act_order', False),
                    gptq_sym=getattr(args, 'gptq_sym', False),
                    gptq2bit_prune_search_enable=getattr(args, 'gptq2bit_prune_search_enable', False),
                    gptq2bit_prune_m_candidates=getattr(args, 'gptq2bit_prune_m_candidates', list(range(0, 97, 4))),
                    gptq3bit_prune_search_enable=getattr(args, 'gptq3bit_prune_search_enable', False),
                    gptq3bit_prune_m_candidates=getattr(args, 'gptq3bit_prune_m_candidates', list(range(0, 97, 4))),
                    std_gptq_enable=getattr(args, 'std_gptq_enable', False),
                    std_gptq_static_groups=getattr(args, 'std_gptq_static_groups', True),
                    block_gptq_enable=getattr(args, 'std_gptq_enable', False),
                    block_gptq_static_groups=getattr(args, 'std_gptq_static_groups', True),
                    post_gptq_refill_enable=getattr(args, 'post_gptq_refill_enable', False),
                    post_gptq_refill_row_ratio=getattr(args, 'post_gptq_refill_row_ratio', 0.2),
                    post_gptq_refill_pos_ratio=getattr(args, 'post_gptq_refill_pos_ratio', 0.2),
                    in_gptq_refill_enable=getattr(args, 'in_gptq_refill_enable', False),
                    in_gptq_refill_row_ratio=getattr(args, 'in_gptq_refill_row_ratio', 0.2),
                    in_gptq_refill_pos_ratio=getattr(args, 'in_gptq_refill_pos_ratio', 0.2),
                    search_trial_mode=False,
                )
    
                if svd_obr_scheme == 'search4_linked':
                    candidate_settings = ((1, 0.3), (2, 0.2), (3, 0.1), (4, 0.0))
                    candidate_trials = []
                    best_trial = None
                    for cand_stage, cand_ratio in candidate_settings:
                        cand_kwargs = dict(fq_kwargs)
                        cand_kwargs.update({
                            'svd_2bit_stages': int(cand_stage),
                            'svd_2bit_obr_twogroup_enable': True,
                            'svd_2bit_obr_twogroup_4bit_ratio': float(cand_ratio),
                            'svd_2bit_obr_twogroup_adaptive_enable': False,
                            'svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio': False,
                            'svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor': False,
                            'svd_obr_interblock_gptq_propagation_disable': True,
                            'search_trial_mode': True,
                        })
                        cand_info = gptq[name].fasterquant(**cand_kwargs)
                        cand_score = float(cand_info.get('activation_error', float('inf'))) if isinstance(cand_info, dict) else float('inf')
                        candidate_trials.append((int(cand_stage), float(cand_ratio), cand_score))
                        if (best_trial is None) or (cand_score < best_trial[0]):
                            best_trial = (cand_score, int(cand_stage), float(cand_ratio))
    
                    logger.info(
                        "[Scheme search4_linked] candidates: " + ", ".join(
                            [f"stage={s},ratio_4bit={r:.1f},act_qerror={err:.4f}" for (s, r, err) in candidate_trials]
                        )
                    )
    
                    if best_trial is None:
                        raise RuntimeError(f'No valid candidate found for scheme=search4_linked at layer {layer_key}.')
    
                    best_score, best_stage, best_ratio = best_trial
                    logger.info(
                        f"[Scheme search4_linked] selected stage={best_stage}, ratio_4bit={best_ratio:.2f}, "
                        f"min_activation_qerror={best_score:.4f}"
                    )
    
                    final_kwargs = dict(fq_kwargs)
                    final_kwargs.update({
                        'svd_2bit_stages': int(best_stage),
                        'svd_2bit_obr_twogroup_enable': True,
                        'svd_2bit_obr_twogroup_4bit_ratio': float(best_ratio),
                        'svd_2bit_obr_twogroup_adaptive_enable': False,
                        'svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio': False,
                        'svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor': False,
                        'svd_obr_interblock_gptq_propagation_disable': True,
                        'search_trial_mode': False,
                    })
                    info = gptq[name].fasterquant(**final_kwargs)
                    if isinstance(info, dict):
                        info['svd_scheme_selection'] = {
                            'scheme': 'search4_linked',
                            'selected_stage': int(best_stage),
                            'selected_ratio_4bit': float(best_ratio),
                            'selected_activation_error': float(best_score),
                            'candidates': [
                                {
                                    'stage': int(s),
                                    'ratio_4bit': float(r),
                                    'activation_error': float(err),
                                }
                                for (s, r, err) in candidate_trials
                            ],
                        }
                else:
                    info = gptq[name].fasterquant(**fq_kwargs)

                if getattr(args, 'svd_energy_stats_enable', False):
                    trace = info.get("svd_energy_trace", []) if isinstance(info, dict) else []
                    if trace:
                        layer_numel = int(subset[name].weight.numel())
                        layer_full_name = f"{i}.{name}"
                        for event in trace:
                            if not isinstance(event, dict):
                                continue
                            record = dict(event)
                            record["layer"] = layer_full_name
                            record["param_count"] = layer_numel
                            svd_energy_events.append(record)

                            stage_name = str(record.get("stage", "unknown"))
                            probe_rank = int(record.get("probe_rank", getattr(args, 'svd_energy_probe_rank', 128)))
                            e_probe = float(record.get("energy_at_probe_rank", 0.0))
                            msg = (
                                f"SVD energy trace layer={layer_full_name}, stage={stage_name}, "
                                f"E({probe_rank})={e_probe:.6f}"
                            )
                            rank_map = record.get("rank_for_threshold", {})
                            if isinstance(rank_map, dict):
                                for tau_key, rank_tau in rank_map.items():
                                    try:
                                        tau_val = float(tau_key)
                                    except (TypeError, ValueError):
                                        continue
                                    msg += f", r@{tau_val:.2f}={int(rank_tau)}"
                            logger.info(msg)

                bit_stats = info.get("bit_stats") if isinstance(info, dict) else None
                if isinstance(bit_stats, dict):
                    layer_numel = int(subset[name].weight.numel())
                    layer_full_name = f"{i}.{name}"
                    bit_record = dict(bit_stats)
                    bit_record["layer"] = layer_full_name
                    bit_record["param_count"] = layer_numel
                    svd_bit_stats_events.append(bit_record)
                    logger.info(
                        f"SVD bit stats layer={layer_full_name}, "
                        f"avg_weight_bits={float(bit_record.get('avg_weight_bits', 0.0)):.6f}, "
                        f"avg_total_bits={float(bit_record.get('avg_total_bits', 0.0)):.6f}, "
                        f"residual_avg_bits={float(bit_record.get('residual_avg_data_bits', 0.0)):.6f}"
                    )

                if isinstance(info, dict) and (info.get("layer_avg_weight_bits") is not None):
                    layer_numel = int(subset[name].weight.numel())
                    layer_full_name = f"{i}.{name}"
                    layer_record = {
                        "layer": layer_full_name,
                        "param_count": layer_numel,
                        "avg_weight_bits": float(info.get("layer_avg_weight_bits", 0.0)),
                        "residual_avg_bits": float(info.get("layer_avg_residual_bits", 0.0) or 0.0),
                        "lowrank_avg_bits": float(info.get("layer_avg_lowrank_bits", 0.0) or 0.0),
                        "residual_pruned_ratio": float(info.get("layer_residual_pruned_ratio", 0.0) or 0.0),
                    }
                    layer_avg_weight_bits_events.append(layer_record)
                    logger.info(
                        f"Layer avg weight bits layer={layer_full_name}, "
                        f"avg_weight_bits={layer_record['avg_weight_bits']:.6f}, "
                        f"residual_avg_bits={layer_record['residual_avg_bits']:.6f}, "
                        f"lowrank_avg_bits={layer_record['lowrank_avg_bits']:.6f}, "
                        f"residual_pruned_ratio={layer_record['residual_pruned_ratio'] * 100:.2f}%"
                    )
                    if bool(info.get("post_gptq_refill_enabled", False)):
                        refill_nonzero = float(info.get("post_gptq_refill_nonzero_count", 0.0) or 0.0)
                        refill_candidates = float(info.get("post_gptq_refill_candidate_count", 0.0) or 0.0)
                        refill_ratio = float(info.get("post_gptq_refill_ratio", 0.0) or 0.0)
                        refill_total_pruned = float(info.get("post_gptq_refill_total_pruned_count", 0.0) or 0.0)
                        refill_selected_rows = float(info.get("post_gptq_refill_selected_row_count", 0.0) or 0.0)
                        refill_err_before = float(info.get("post_gptq_refill_error_before", 0.0) or 0.0)
                        refill_err_after = float(info.get("post_gptq_refill_error_after", 0.0) or 0.0)
                        logger.info(
                            f"Layer post-gptq refill layer={layer_full_name}, "
                            f"refilled_nonzero={int(refill_nonzero)}/{int(refill_candidates)} "
                            f"({refill_ratio * 100:.2f}%), "
                            f"selected_rows={int(refill_selected_rows)}, "
                            f"total_pruned={int(refill_total_pruned)}, "
                            f"row_abs_error_before={refill_err_before:.6e}, "
                            f"row_abs_error_after={refill_err_after:.6e}"
                        )
    
                gptq[name].free()

            if cascade_analysis_enabled and i == 0:
                _export_cascade_analysis_excel(
                    layer_rows_map=cascade_analysis_rows,
                    output_path=CASCADE_ANALYSIS_OUTPUT,
                )
                print("Mission Accomplished", flush=True)
                os._exit(0)

            if residual_dist_analysis_enabled and i == 0:
                excel_path = os.path.join(residual_dist_output_dir, RESIDUAL_DIST_ANALYSIS_OUTPUT)
                _export_residual_distribution_excel(
                    layer_results=residual_dist_layer_results,
                    output_path=excel_path,
                )
                print("Mission Accomplished", flush=True)
                os._exit(0)

        for j in range(args.nsamples):
            outs[j] = forward_layer_with_position(layer, inps[j].unsqueeze(0), attention_mask, position_ids)
        layers[i] = layer.cpu()
        del layer

        inps, outs = outs, inps

    if getattr(args, 'svd_energy_stats_enable', False):
        if len(svd_energy_events) == 0:
            logger.info("SVD energy summary: no SVD decomposition events were recorded.")
        else:
            logger.info("=" * 80)
            logger.info(f"SVD energy summary: total events={len(svd_energy_events)}")

            param_weights = [float(x.get("param_count", 1.0)) for x in svd_energy_events]
            total_weight = sum(param_weights)

            energy_vals = [float(x.get("energy_at_probe_rank", 0.0)) for x in svd_energy_events]
            mean_energy = sum(energy_vals) / len(energy_vals)
            if total_weight > 0:
                weighted_energy = sum(v * w for v, w in zip(energy_vals, param_weights)) / total_weight
            else:
                weighted_energy = 0.0

            used_ranks = [float(x.get("used_rank", 0.0)) for x in svd_energy_events]
            mean_used_rank = sum(used_ranks) / len(used_ranks)
            if total_weight > 0:
                weighted_used_rank = sum(v * w for v, w in zip(used_ranks, param_weights)) / total_weight
            else:
                weighted_used_rank = 0.0

            probe_rank = int(getattr(args, 'svd_energy_probe_rank', 128))
            logger.info(
                f"  E({probe_rank}) mean={mean_energy:.6f}, weighted_by_params={weighted_energy:.6f}"
            )
            logger.info(
                f"  used_rank mean={mean_used_rank:.2f}, weighted_by_params={weighted_used_rank:.2f}"
            )

            thresholds = [float(t) for t in getattr(args, 'svd_energy_thresholds', [0.9, 0.92])]
            for tau in thresholds:
                tau_ranks = []
                tau_weighted_sum = 0.0
                tau_weighted_den = 0.0

                for event in svd_energy_events:
                    rank_map = event.get("rank_for_threshold", {})
                    if not isinstance(rank_map, dict):
                        continue

                    rank_tau = None
                    for tau_key, rank_val in rank_map.items():
                        try:
                            tau_key_val = float(tau_key)
                        except (TypeError, ValueError):
                            continue
                        if abs(tau_key_val - tau) <= 1e-6:
                            rank_tau = float(rank_val)
                            break

                    if rank_tau is None:
                        continue

                    tau_ranks.append(rank_tau)
                    w = float(event.get("param_count", 1.0))
                    tau_weighted_sum += rank_tau * w
                    tau_weighted_den += w

                if len(tau_ranks) == 0:
                    logger.info(f"  tau={tau:.2f}: no rank stats collected")
                    continue

                mean_tau_rank = sum(tau_ranks) / len(tau_ranks)
                weighted_tau_rank = tau_weighted_sum / tau_weighted_den if tau_weighted_den > 0 else 0.0
                logger.info(
                    f"  tau={tau:.2f}: required_rank mean={mean_tau_rank:.2f}, "
                    f"weighted_by_params={weighted_tau_rank:.2f}"
                )

            logger.info("=" * 80)

    if len(svd_bit_stats_events) == 0:
        if getattr(args, 'svd_bit_stats_enable', False):
            logger.info("SVD bit-stats summary: no bit stats records were collected.")
    else:
        logger.info("=" * 80)
        logger.info(f"SVD bit-stats summary: total layers={len(svd_bit_stats_events)}")

        param_weights = [float(x.get("param_count", 1.0)) for x in svd_bit_stats_events]
        total_weight = sum(param_weights)

        def _weighted_avg(key):
            vals = [float(x.get(key, 0.0)) for x in svd_bit_stats_events]
            if total_weight <= 0:
                return sum(vals) / max(len(vals), 1)
            return sum(v * w for v, w in zip(vals, param_weights)) / total_weight

        avg_weight_bits = _weighted_avg("avg_weight_bits")
        avg_total_bits = _weighted_avg("avg_total_bits")
        avg_meta_bits = _weighted_avg("avg_meta_overhead_bits")
        avg_residual_bits = _weighted_avg("residual_avg_data_bits")
        avg_residual_total_bits = _weighted_avg("residual_avg_total_bits")
        avg_psal = _weighted_avg("p_sal_real")

        logger.info(
            f"  avg_weight_bits(data_only)={avg_weight_bits:.6f}, "
            f"avg_total_bits(data+meta)={avg_total_bits:.6f}, "
            f"avg_meta_overhead_bits={avg_meta_bits:.6f}"
        )
        logger.info(
            f"  residual_avg_bits(data_only)={avg_residual_bits:.6f}, "
            f"residual_avg_total_bits(data+meta)={avg_residual_total_bits:.6f}, "
            f"avg_p_sal={avg_psal * 100:.2f}%"
        )
        logger.info("=" * 80)

    if len(layer_avg_weight_bits_events) == 0:
        logger.info("Layer avg weight bits summary: no records were collected.")
        if bool(getattr(args, "svd_enable", False)) and bool(getattr(args, "std_gptq_enable", False)):
            logger.info(
                "Layer avg weight bits summary note: current SVD+std_gptq path "
                "does not emit per-layer avg-bit records yet."
            )
    else:
        logger.info("=" * 80)
        logger.info(f"Layer avg weight bits summary: total layers={len(layer_avg_weight_bits_events)}")
        total_params = sum(float(x.get("param_count", 0.0)) for x in layer_avg_weight_bits_events)
        if total_params <= 0:
            total_params = float(len(layer_avg_weight_bits_events))
        weighted_avg_weight_bits = sum(
            float(x.get("avg_weight_bits", 0.0)) * float(x.get("param_count", 1.0))
            for x in layer_avg_weight_bits_events
        ) / total_params
        weighted_residual_bits = sum(
            float(x.get("residual_avg_bits", 0.0)) * float(x.get("param_count", 1.0))
            for x in layer_avg_weight_bits_events
        ) / total_params
        weighted_lowrank_bits = sum(
            float(x.get("lowrank_avg_bits", 0.0)) * float(x.get("param_count", 1.0))
            for x in layer_avg_weight_bits_events
        ) / total_params
        weighted_pruned_ratio = sum(
            float(x.get("residual_pruned_ratio", 0.0)) * float(x.get("param_count", 1.0))
            for x in layer_avg_weight_bits_events
        ) / total_params
        logger.info(
            f"  [Global Avg Weight Bits] avg_weight_bits={weighted_avg_weight_bits:.6f}, "
            f"residual_avg_bits={weighted_residual_bits:.6f}, "
            f"lowrank_avg_bits={weighted_lowrank_bits:.6f}, "
            f"residual_pruned_ratio={weighted_pruned_ratio * 100:.2f}%"
        )
        logger.info(
            f"  weighted_avg_weight_bits={weighted_avg_weight_bits:.6f}, "
            f"weighted_residual_avg_bits={weighted_residual_bits:.6f}, "
            f"weighted_lowrank_avg_bits={weighted_lowrank_bits:.6f}, "
            f"weighted_residual_pruned_ratio={weighted_pruned_ratio * 100:.2f}%"
        )
        logger.info("=" * 80)

    model.config.use_cache = use_cache


if __name__ == "__main__":
    import argparse
    from datautils import *

    def list_of_ints(arg):
        return list(map(int, arg.split(',')))
    
    def list_of_floats(arg):
        return list(map(float, arg.split(',')))

    parser = argparse.ArgumentParser()

    parser.add_argument(
        "model", type=str, help="model to load; for example `huggyllama/llama-7b`."
    )
    parser.add_argument(
        "dataset",
        type=str,
        choices=["wikitext2", "ptb", "c4"],
        help="Where to extract calibration data from.",
    )
    parser.add_argument(
        "low_quant_method",
        type=str,
        choices=["xnor", "sign", "no", "2bit", "3bit", "4bit", "prune", "braq"],
        help="quantization method; `xnor` is the method using XNOR to adapt hardware calculation; `prune` is the method used in sparseGPTQ; braq is the method used in BiLLM; `3bit` uses 3-bit quantization with GPTQ (skips all masks)",
    )
    parser.add_argument(
        "--seed", type=int, default=0, help="Seed for sampling the calibration data."
    )
    parser.add_argument(
        "--nsamples", type=int, default=128, help="Number of calibration data samples."
    )
    parser.add_argument(
        "--percdamp",
        type=float,
        default=0.01,
        help="Percent of the average Hessian diagonal to use for dampening.",
    )
    parser.add_argument(
        "--blocksize",
        type=int,
        default=128,
        help="Blocksize to use for adaptive mask selection.",
    )
    parser.add_argument(
        "--salient_metric",
        type=str,
        default="magnitude",
        choices=["magnitude", "hessian", "wanda"],
        help="Metric for determining salient weights: 'magnitude' (weight magnitude), 'hessian' (Hessian-based), or 'wanda' (weight * activation norm).",
    )
    parser.add_argument(
        "--device",
        type=str,
        default="cuda:0",
        help="set the device to use for quantization.",
    )
    parser.add_argument(
        "--disable_gptq",
        action="store_true",
        help="disable GPTQ for quantization.",
    )
    parser.add_argument(
        "--std_gptq_enable",
        action="store_true",
        default=False,
        help=(
            "Enable std_gptq path: standard low-bit (2bit/3bit/4bit) GPTQ with column-wise in-block "
            "compensation and inter-block compensation."
        ),
    )
    parser.add_argument(
        "--std_gptq_static_groups",
        dest="std_gptq_static_groups",
        action="store_true",
        default=True,
        help="Use static per-group qparams in std_gptq when groupsize is enabled (default: enabled).",
    )
    parser.add_argument(
        "--std_gptq_dynamic_groups",
        dest="std_gptq_static_groups",
        action="store_false",
        help="Use dynamic per-group qparams in std_gptq (recompute qparams during block updates).",
    )
    parser.add_argument(
        "--block_gptq_enable",
        dest="std_gptq_enable",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--block_gptq_static_groups",
        dest="std_gptq_static_groups",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--block_gptq_dynamic_groups",
        dest="std_gptq_static_groups",
        action="store_false",
        help=argparse.SUPPRESS,
    )
    parser.add_argument(
        "--pure_gptq",
        action="store_true",
        default=False,
        help="Deprecated alias. Use --std_gptq_enable instead.",
    )
    parser.add_argument(
        "--post_gptq_refill_enable",
        action="store_true",
        default=False,
        help="Enable lightweight post-GPTQ refill on row-pruned positions (std_gptq rowprune_2pass path only).",
    )
    parser.add_argument(
        "--post_gptq_refill_row_ratio",
        type=float,
        default=0.2,
        help="Top row ratio (by row output error proxy) used by post-GPTQ refill.",
    )
    parser.add_argument(
        "--post_gptq_refill_pos_ratio",
        type=float,
        default=0.2,
        help="Top pruned-position ratio per selected row used by post-GPTQ refill.",
    )
    parser.add_argument(
        "--in_gptq_refill_enable",
        action="store_true",
        default=False,
        help=(
            "Enable in-GPTQ residual refill on row-pruned positions. "
            "Refill is applied inside GPTQ column updates so refill errors are compensated by Hessian propagation. "
            "Mutually exclusive with --post_gptq_refill_enable."
        ),
    )
    parser.add_argument(
        "--in_gptq_refill_row_ratio",
        type=float,
        default=0.2,
        help="Top row ratio (by row importance proxy) used by in-GPTQ residual refill.",
    )
    parser.add_argument(
        "--in_gptq_refill_pos_ratio",
        type=float,
        default=0.2,
        help="Top pruned-position ratio per selected row used by in-GPTQ residual refill.",
    )
    parser.add_argument(
        "--gptq_groupsize",
        type=int,
        default=-1,
        help="Groupsize for std_gptq qparam estimation (-1 means fallback to quantizer groupsize).",
    )
    parser.add_argument(
        "--gptq_static_groups",
        dest="gptq_static_groups",
        action="store_true",
        default=False,
        help="Legacy compatibility flag for static per-group qparams (deprecated; prefer --std_gptq_static_groups).",
    )
    parser.add_argument(
        "--gptq_dynamic_groups",
        dest="gptq_static_groups",
        action="store_false",
        help="Legacy compatibility flag for dynamic per-group qparams (deprecated; prefer --std_gptq_dynamic_groups).",
    )
    parser.add_argument(
        "--gptq_act_order",
        action="store_true",
        default=False,
        help="Enable standard GPTQ act-order heuristic in std_gptq mode.",
    )
    parser.add_argument(
        "--gptq_sym",
        action="store_true",
        default=False,
        help="Use symmetric quantization for std_gptq quantizer params.",
    )
    parser.add_argument(
        "--gptq_true_sequential",
        action="store_true",
        default=False,
        help="Use true-sequential submodule grouping in std_gptq mode.",
    )
    parser.add_argument(
        "--gptq2bit_prune_search_enable",
        action="store_true",
        default=False,
        help="Deprecated. This legacy pure-gptq prune-search path has been removed.",
    )
    parser.add_argument(
        "--gptq2bit_prune_m_candidates",
        type=list_of_ints,
        default=list(range(0, 97, 4)),
        help="Comma-separated row-prune m candidates for --gptq2bit_prune_search_enable, e.g. 0,4,8,...,96.",
    )
    parser.add_argument(
        "--gptq3bit_prune_search_enable",
        action="store_true",
        default=False,
        help="Deprecated. This legacy pure-gptq prune-search path has been removed.",
    )
    parser.add_argument(
        "--gptq3bit_prune_m_candidates",
        type=list_of_ints,
        default=list(range(0, 97, 4)),
        help="Comma-separated row-prune m candidates for --gptq3bit_prune_search_enable, e.g. 0,4,8,...,96.",
    )
    parser.add_argument(
        "--disable_iterative_2bit",
        action="store_true",
        default=False,
        help="Disable masked iterative affine quantization in legacy non-std_gptq paths. "
             "In std_gptq 2bit/3bit/4bit, iterative affine is mandatory and this flag is rejected.",
    )
    parser.add_argument(
        "--iterative_2bit_iters",
        type=int,
        default=5,
        help="Minimum alternating iterations before convergence-stop in masked iterative affine quantization (effective min is 5).",
    )
    parser.add_argument(
        "--iterative_2bit_fixed_iters",
        action="store_true",
        default=False,
        help="Run masked iterative affine quantization with a fixed number of iterations (no early-stop).",
    )
    parser.add_argument(
        "--iterative_2bit_eps",
        type=float,
        default=1e-8,
        help="Numerical epsilon for masked iterative affine quantization.",
    )
    parser.add_argument(
        "--iterative_2bit_disable_offset_update",
        action="store_true",
        default=False,
        help="Disable iterative offset update in masked low-bit quantization (optimize scale only).",
    )
    parser.add_argument(
        "--iterative_2bit_print_mse",
        action="store_true",
        default=False,
        help="Print masked MSE for each alternating iteration in iterative low-bit quantization.",
    )
    parser.add_argument(
        "--svd256_then_pure_gptq2bit",
        action="store_true",
        default=False,
        help="Preset mode: run standard truncated SVD(rank=256), then quantize residual with std_gptq 2bit.",
    )
    parser.add_argument(
        "--svd256_then_pure_gptq3bit",
        action="store_true",
        default=False,
        help="Preset mode: run standard truncated SVD(rank=256), then quantize residual with std_gptq 3bit.",
    )
    parser.add_argument(
        "--minlayer", type=int, default=-1, help="Quant all layers with id >= this."
    )
    parser.add_argument(
        "--maxlayer", type=int, default=1000, help="Quant all layers with id < this."
    )
    parser.add_argument(
        "--quant_only",
        type=str,
        default="",
        help="Quant only layers that contain this text.",
    )
    parser.add_argument("--invert", action="store_true", help="Invert subset.")
    parser.add_argument(
        "--save",
        action="store_true",
    )
    parser.add_argument(
        "--log_wandb", action="store_true", help="Whether to log to wandb."
    )
    parser.add_argument(
        "--svd_enable",
        action="store_true",
        default=False,
        help="Enable SVD decomposition for quantization.",
    )
    parser.add_argument(
        "--svd_rank",
        type=int,
        default=8,
        help="Rank for SVD decomposition (default: 8).",
    )
    parser.add_argument(
        "--svd_auto_rank_deltae_stop",
        action="store_true",
        default=False,
        help="Single-stage SVD rank auto-search mode: start rank=16, step=16, stop when Delta(normalized activation error) < 0.05; residual uses full-block order=2 binary approximation.",
    )
    parser.add_argument(
        "--svd_rank_by_vh_range_sigma",
        action="store_true",
        default=False,
        help="Deprecated. |range(vh)|*|sigma| component selection has been removed; this flag is ignored.",
    )
    parser.add_argument(
        "--svd_rank_by_vh_range_sigma_topk",
        type=int,
        default=32,
        help="Deprecated. Kept only for backward compatibility and ignored.",
    )
    parser.add_argument(
        "--svd_energy_stats_enable",
        action="store_true",
        default=False,
        help="Enable SVD energy statistics logging (E(probe_rank) and required rank for energy thresholds).",
    )
    parser.add_argument(
        "--svd_energy_probe_rank",
        type=int,
        default=128,
        help="Probe rank used in E(probe_rank) logging when --svd_energy_stats_enable is set (default: 128).",
    )
    parser.add_argument(
        "--svd_energy_thresholds",
        type=list_of_floats,
        default=[0.90, 0.92],
        help="Comma-separated energy thresholds used to report required rank, e.g. 0.9,0.92.",
    )
    parser.add_argument(
        "--svd_bit_stats_enable",
        action="store_true",
        default=False,
        help="Enable per-layer and model-level bit accounting logs (data bits + scale/zp overhead).",
    )
    parser.add_argument(
        "--svd_bit_stats_scale_bits",
        type=float,
        default=16,
        help="Bit width used for each stored scale in bit accounting (default: 16).",
    )
    parser.add_argument(
        "--svd_bit_stats_zp_bits",
        type=float,
        default=8,
        help="Bit width used for each stored zero-point in bit accounting (default: 8).",
    )
    parser.add_argument(
        "--svd_bit_stats_binary_data_bits",
        type=float,
        default=2.0,
        help="Effective data-bit width used for non-salient binary-residual branch in accounting formula (default: 2.0).",
    )
    parser.add_argument(
        "--svd_lwc_enable",
        action="store_true",
        default=False,
        help="Enable per-block LWC clip-ratio search on SVD residual before full-block standard 2bit per-channel quantization. Selection objective is block weight-loss sum.",
    )
    parser.add_argument(
        "--svd_lwc_candidates",
        type=list_of_floats,
        default=[1.0, 0.999, 0.995, 0.99, 0.98],
        help="Comma-separated LWC clip-ratio candidates (percentiles over |R_block|), e.g. 1.0,0.999,0.995,0.99,0.98.",
    )
    parser.add_argument(
        "--svd_row_prune_search_enable",
        action="store_true",
        default=False,
        help="Enable per-block row-prune m-search before binary residual approximation (order configurable), objective L(m)=sum((R_b-Q_b)^2) with original unpruned R_b.",
    )
    parser.add_argument(
        "--svd_row_prune_m_candidates",
        type=list_of_ints,
        default=[0, 4, 8, 16, 32, 64],
        help="Comma-separated row-prune m candidates per row, e.g. 0,4,8,16,32,64.",
    )
    row_prune_score_group = parser.add_mutually_exclusive_group()
    row_prune_score_group.add_argument(
        "--svd_row_prune_score_magnitude",
        dest="svd_row_prune_score_metric",
        action="store_const",
        const="magnitude",
        help="Use |w| score for row-prune mask search (prune smallest |w| per row).",
    )
    row_prune_score_group.add_argument(
        "--svd_row_prune_score_wanda",
        dest="svd_row_prune_score_metric",
        action="store_const",
        const="wanda",
        help="Use WANDA score |w|*sqrt(E[x^2]) for row-prune mask search (prune smallest score per row).",
    )
    parser.set_defaults(svd_row_prune_score_metric="magnitude")
    parser.add_argument(
        "--svd_row_prune_binary_order",
        type=int,
        default=2,
        help="Binary residual approximation order used in row-prune search metric and final residual quantization (1/2/3).",
    )
    parser.add_argument(
        "--svd_row_prune_quant_scheme",
        type=str,
        default="binary",
        choices=["binary", "std2bit", "std3bit", "std4bit"],
        help="Quantization scheme for row-prune search metric and final residual quantization: "
        "binary / std2bit / std3bit / std4bit.",
    )
    parser.add_argument(
        "--svd_row_prune_clip_search_enable",
        action="store_true",
        default=False,
        help="Enable second-stage clip search AFTER best row-prune m is selected (m-search first, then clip-search).",
    )
    parser.add_argument(
        "--svd_row_prune_clip_candidates",
        type=list_of_floats,
        default=[1.0, 0.999, 0.995, 0.99, 0.98, 0.96],
        help="Comma-separated clip percentile candidates over kept |R_block| after row-prune, e.g. 1.0,0.999,0.995,0.99,0.98,0.96.",
    )
    parser.add_argument(
        "--svd_row_prune_clip_min_value",
        type=float,
        default=1e-8,
        help="Minimum positive clip value used in row-prune clip search (default: 1e-8).",
    )
    parser.add_argument(
        "--svd_row_prune_act_topk",
        type=int,
        default=3,
        help="Top-K m candidates (ranked by weight loss) to rerank by block-Hessian activation score after one-shot RTN refill on pruned entries; <=0 disables rerank.",
    )
    parser.add_argument(
        "--svd_row_prune_offline_refill_enable",
        action="store_true",
        default=False,
        help="Enable offline post-quant refill on row-pruned positions using full-row activation objective e^T H e.",
    )
    parser.add_argument(
        "--svd_row_prune_offline_refill_sweeps",
        type=int,
        default=1,
        help="Coordinate-descent sweep count for row-prune offline refill (default: 1).",
    )
    parser.add_argument(
        "--svd_row_prune_offline_refill_max_positions",
        type=int,
        default=-1,
        help="Per-row max pruned positions for refill updates; <=0 means all pruned positions.",
    )
    parser.add_argument(
        "--svd_row_prune_offline_refill_exclude_zero",
        action="store_false",
        dest="svd_row_prune_offline_refill_include_zero",
        default=True,
        help="Exclude zero from refill candidate set (default behavior includes zero).",
    )

    parser.add_argument(
        "--adaptive_rank",
        action="store_true",
        default=False,
        help="Enable adaptive rank allocation based on layer sensitivity. Requires --svd_enable.",
    )
    parser.add_argument(
        "--avg_rank",
        type=int,
        default=256,
        help="Average rank per layer for adaptive allocation. Total budget = num_layers * avg_rank (default: 256).",
    )
    parser.add_argument(
        "--min_rank",
        type=int,
        default=128,
        help="Minimum rank for adaptive allocation (default: 128).",
    )
    parser.add_argument(
        "--max_rank",
        type=int,
        default=512,
        help="Maximum rank for adaptive allocation (default: 512).",
    )
    parser.add_argument(
        "--block_greedy_rank",
        action="store_true",
        default=False,
        help="Enable block-wise greedy rank allocation based on entropy reduction. Requires --adaptive_rank and --svd_enable.",
    )
    parser.add_argument(
        "--block_greedy_base_rank",
        type=int,
        default=128,
        help="Base rank for block-wise greedy allocation (default: 128).",
    )
    parser.add_argument(
        "--block_greedy_max_rank",
        type=int,
        default=512,
        help="Maximum rank for SVD pre-computation in block-wise greedy allocation (default: 512).",
    )
    parser.add_argument(
        "--block_greedy_step_size",
        type=int,
        default=32,
        help="Step size for block-wise greedy allocation (default: 32).",
    )
    parser.add_argument(
        "--use_activation_entropy",
        action="store_true",
        default=False,
        help="Use stable rank-based scoring instead of energy gain. Formula: Score = ΔH_l × log10(stable_rank). Requires --block_greedy_rank.",
    )
    parser.add_argument(
        "--activation_entropy_beta",
        type=float,
        default=2.0,
        help="Beta coefficient for activation entropy in scoring formula. Used when --use_activation_entropy is enabled (default: 2.0).",
    )
    parser.add_argument(
        "--disable_two_level_allocation",
        action="store_true",
        default=False,
        help="Disable two-level allocation strategy (block-level + layer-level). Only use single-level allocation within each block. Requires --block_greedy_rank.",
    )
    parser.add_argument(
        "--svd_drop_residual",
        action="store_true",
        default=False,
        help="Drop residual R after SVD decomposition, only keep low-rank L in FP16. Requires --svd_enable.",
    )
    parser.add_argument(
        "--use_bass",
        action="store_true",
        default=False,
        help="Deprecated. BASS path has been removed; this flag is ignored.",
    )
    parser.add_argument(
        "--bass_pool_factor",
        type=int,
        default=None,
        help="Deprecated. Kept only for backward compatibility and ignored.",
    )
    parser.add_argument(
        "--bass_batch_size",
        type=int,
        default=None,
        help="Deprecated. Kept only for backward compatibility and ignored.",
    )
    parser.add_argument(
        "--bass_selection_mode",
        type=str,
        default="greedy",
        choices=["greedy", "proxy"],
        help="Deprecated. Kept only for backward compatibility and ignored.",
    )
    parser.add_argument(
        "--bass_proxy_metric",
        type=str,
        default="linf",
        choices=["linf", "energy"],
        help="Deprecated. Kept only for backward compatibility and ignored.",
    )
    parser.add_argument(
        "--linf_sigma_weight",
        type=float,
        default=1.0,
        help="Deprecated. Kept only for backward compatibility and ignored.",
    )
    parser.add_argument(
        "--linf_vh_weight",
        type=float,
        default=1.0,
        help="Deprecated. Kept only for backward compatibility and ignored.",
    )
    parser.add_argument(
        "--swbe_ref_rank",
        type=int,
        default=256,
        help="Deprecated. Kept only for backward compatibility and ignored.",
    )
    parser.add_argument(
        "--swbe_total_pool",
        type=int,
        default=None,
        help="Deprecated. Kept only for backward compatibility and ignored.",
    )
    parser.add_argument(
        "--svd_num_iters",
        type=int,
        default=1,
        help="Number of SVD-BiLLM iterative optimization iterations (default: 1). Total iterations = svd_num_iters (first iteration + additional iterations).",
    )
    parser.add_argument(
        "--svd_lowrank_fp16",
        action="store_true",
        default=False,
        help="Keep the SVD low-rank branch in FP16 (no low-rank 4bit quantization). Useful for diagnostics like L_fp16 + R_2bit.",
    )
    parser.add_argument(
        "--svd_2bit_stages",
        type=int,
        default=2,
        help="Number of compensated SVD stages used when low_quant_method=2bit and --svd_enable is set (default: 2). Each stage's low-rank branch uses 4-bit by default (or FP16 with --svd_lowrank_fp16); final residual uses GPTQ 2-bit by default (or OBR when --obr is enabled).",
    )
    parser.add_argument(
        "--svd_2bit_extra_8bit_stage_enable",
        action="store_true",
        default=False,
        help="Enable one additional compensated SVD stage whose low-rank branch is quantized to 8-bit per-channel. This stage runs after --svd_2bit_stages.",
    )
    parser.add_argument(
        "--svd_2bit_extra_8bit_stage_rank",
        type=int,
        default=64,
        help="Rank of the optional extra 8-bit low-rank stage in SVD+2bit mode (default: 64).",
    )
    parser.add_argument(
        "--svd_binary_refit_enable",
        action="store_true",
        default=False,
        help="Enable single-stage SVD binary-residual refit mode: quantize residual with order=1 (no mask), refit low-rank factors A,B with fixed residual, then quantize low-rank branch to 4bit.",
    )
    parser.add_argument(
        "--svd_binary_refit_epochs",
        type=int,
        default=5,
        help="Total epochs for low-rank refit in svd_binary_refit mode (default: 5).",
    )
    parser.add_argument(
        "--svd_binary_refit_fp_epochs",
        type=int,
        default=3,
        help="Deprecated compatibility option in svd_binary_refit mode; training is full-FP and 4bit is applied only after refit.",
    )
    parser.add_argument(
        "--svd_binary_refit_lr",
        type=float,
        default=1e-3,
        help="Initial learning rate for low-rank refit in svd_binary_refit mode (default: 1e-3).",
    )
    parser.add_argument(
        "--svd_binary_refit_min_lr",
        type=float,
        default=1e-4,
        help="Final learning rate for low-rank refit schedule in svd_binary_refit mode (default: 1e-4).",
    )
    parser.add_argument(
        "--svd_binary_refit_weight_decay",
        type=float,
        default=1e-5,
        help="Weight decay for AdamW in svd_binary_refit mode (default: 1e-5).",
    )
    parser.add_argument(
        "--svd_binary_refit_grad_clip",
        type=float,
        default=1.0,
        help="Gradient clipping value for low-rank refit in svd_binary_refit mode (default: 1.0).",
    )
    parser.add_argument(
        "--svd_binary_refit_patience",
        type=int,
        default=2,
        help="Early-stop patience for low-rank refit in svd_binary_refit mode (default: 2).",
    )
    parser.add_argument(
        "--svd_binary_refit_max_tokens",
        type=int,
        default=128,
        help="Maximum calibration tokens used to build activation matrix X in svd_binary_refit mode (default: 128).",
    )
    parser.add_argument(
        "--obr",
        action="store_true",
        default=False,
        help="Enable OBR-style residual quantization in SVD+2bit mode. This mode is mutually exclusive with GPTQ and requires --disable_gptq.",
    )
    parser.add_argument(
        "--svd_obr_scheme",
        type=str,
        default="none",
        choices=["none", "adp3_fix01", "search4_linked", "adp4_linked"],
        help="Preset OBR scheme selector: none (default), adp3_fix01 (adaptive 1~3 + fixed salient ratio 0.1), search4_linked (per-layer search over {(1,0.3),(2,0.2),(3,0.1),(4,0.0)} by activation error), adp4_linked (adaptive 1~4 + linked ratio 0.1*(4-n_used)).",
    )
    parser.add_argument(
        "--svd_2bit_salient_3bit_enable",
        action="store_true",
        default=False,
        help="Enable mixed residual quantization in SVD+2bit mode: non-salient columns use binary residual order=2, salient columns use standard 3bit per-channel quantization selected by fixed top-k ratio per block.",
    )
    parser.add_argument(
        "--svd_2bit_salient_4bit_enable",
        action="store_true",
        default=False,
        help="Enable mixed residual quantization in SVD+2bit mode: non-salient columns use binary residual order=2, salient columns use 4bit per-channel (fixed ratio or adaptive search).",
    )
    parser.add_argument(
        "--svd_2bit_obr_cascade_enable",
        action="store_true",
        default=False,
        help="Enable OBR cascade split in SVD+2bit mode: non-salient columns use binary residual order=2, mid-salient columns use standard 3bit per-channel quantization, top-salient columns use 4bit; OBR transfers 2bit->3bit then 3bit->4bit.",
    )
    parser.add_argument(
        "--svd_2bit_obr_cascade_4bit_ratio",
        type=float,
        default=0.1,
        help="Top salient-column ratio assigned to 4bit branch in --svd_2bit_obr_cascade_enable (default: 0.1).",
    )
    parser.add_argument(
        "--svd_2bit_obr_cascade_3bit_ratio",
        type=float,
        default=0.1,
        help="Second salient-column ratio assigned to 3bit branch in --svd_2bit_obr_cascade_enable (default: 0.1).",
    )
    parser.add_argument(
        "--svd_2bit_obr_twogroup_enable",
        action="store_true",
        default=False,
        help="Enable OBR 2-group split in SVD+2bit mode: top-salient columns use 4bit and remaining columns use binary residual order=2; OBR transfers 2bit-group error to 4bit group.",
    )
    parser.add_argument(
        "--svd_2bit_obr_twogroup_4bit_ratio",
        type=float,
        default=0.1,
        help="Top salient-column ratio assigned to 4bit group in --svd_2bit_obr_twogroup_enable (default: 0.1).",
    )
    parser.add_argument(
        "--svd_2bit_obr_twogroup_adaptive_enable",
        action="store_true",
        default=False,
        help="Enable adaptive OBR 2-group mode: run up to max SVD stages and continue while E(128)>=threshold; residual 4bit ratio becomes max(1, max_stages + 1 - used_stages)*base_ratio.",
    )
    parser.add_argument(
        "--svd_2bit_obr_twogroup_adaptive_max_stages",
        type=int,
        default=3,
        help="Maximum SVD stage count used by --svd_2bit_obr_twogroup_adaptive_enable (default: 3).",
    )
    parser.add_argument(
        "--svd_2bit_obr_twogroup_adaptive_e128_threshold",
        type=float,
        default=0.2,
        help="Stage-continue threshold in adaptive twogroup mode: continue next stage when current E(128)>=threshold (default: 0.2).",
    )
    parser.add_argument(
        "--svd_2bit_obr_twogroup_adaptive_base_ratio",
        type=float,
        default=0.1,
        help="Base ratio for adaptive twogroup residual split: ratio_4bit=max(1, max_stages + 1 - used_stages)*base_ratio (default: 0.1).",
    )
    parser.add_argument(
        "--svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio",
        action="store_true",
        default=False,
        help="In adaptive twogroup mode, keep 4bit salient ratio fixed to --svd_2bit_obr_twogroup_4bit_ratio (default 0.1) instead of using stage-count-based ratio.",
    )
    parser.add_argument(
        "--svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor",
        action="store_true",
        default=False,
        help="Use linked-ratio mode without floor in adaptive twogroup: ratio_4bit = base_ratio * max(0, max_stages - used_stages).",
    )
    parser.add_argument(
        "--svd_2bit_obr_4group_enable",
        action="store_true",
        default=False,
        help="Enable OBR 4-group split in SVD+2bit mode: top-salient columns use 4bit, remaining columns are split into three 2bit groups by Wanda/Hessian/magnitude score; OBR transfers errors g3->g2->g1->g0 in each block, and inter-block GPTQ propagation is enabled.",
    )
    parser.add_argument(
        "--svd_2bit_obr_4group_4bit_ratio",
        type=float,
        default=0.1,
        help="Top salient-column ratio assigned to the 4bit group in --svd_2bit_obr_4group_enable (default: 0.1).",
    )
    parser.add_argument(
        "--svd_2bit_obr_4group_2bit_ratio_1",
        type=float,
        default=0.3,
        help="Second group ratio assigned to the first 2bit group in --svd_2bit_obr_4group_enable (default: 0.3).",
    )
    parser.add_argument(
        "--svd_2bit_obr_4group_2bit_ratio_2",
        type=float,
        default=0.3,
        help="Third group ratio assigned to the second 2bit group in --svd_2bit_obr_4group_enable (default: 0.3). Remaining columns form the last 2bit group.",
    )
    parser.add_argument(
        "--svd_2bit_obr_4group_non4bit_scheme",
        type=str,
        default="order2",
        choices=["order2", "std2bit", "ternary", "ttb1"],
        help="Non-4bit quantization scheme for g1/g2/g3 in --svd_2bit_obr_4group_enable: "
             "order2 (binary residual order=2), std2bit (standard 2bit per-channel), "
             "ternary (TWN-style {-1,0,1}), ttb1 (g1/g2 ternary + g3 binary order=1).",
    )
    parser.add_argument(
        "--svd_2bit_obr_ternary_delta_factor",
        type=float,
        default=0.7,
        help="Delta factor in TWN-style ternary quantization: Delta = factor * mean(|W|) (default: 0.7).",
    )
    parser.add_argument(
        "--svd_obr_interblock_gptq_propagation_disable",
        action="store_true",
        default=False,
        help="Disable inter-block Hessian/GPTQ-style error propagation in OBR split modes. If not set, propagation is enabled.",
    )
    parser.add_argument(
        "--svd_obr_cascade_lowrank_8bit_enable",
        action="store_true",
        default=False,
        help="Enable one-shot low-rank correction after R cascade OBR: run 1-stage SVD first, quantize R via 2->3->4 cascade, transfer only 4bit-group error to low-rank branch with closed-form OBR-like projection, then quantize low-rank factors to 8bit.",
    )
    parser.add_argument(
        "--svd_obr_cascade_lowrank_gamma",
        type=float,
        default=1.0,
        help="Scaling factor for 4bit-group residual error transferred to low-rank target in --svd_obr_cascade_lowrank_8bit_enable (default: 1.0).",
    )
    parser.add_argument(
        "--svd_obr_cascade_lowrank_damp",
        type=float,
        default=0.01,
        help="Damping scale used in low-rank projected OBR solve in --svd_obr_cascade_lowrank_8bit_enable (default: 0.01).",
    )
    parser.add_argument(
        "--svd_2bit_salient_3bit_ratio",
        type=float,
        default=0.1,
        help="Salient-column ratio setting in SVD+2bit mixed mode (default: 0.1). For 3bit mode and fixed 4bit mode it is fixed top-k ratio; for adaptive 4bit mode it is max salient ratio.",
    )
    parser.add_argument(
        "--svd_2bit_salient_4bit_alpha",
        type=float,
        default=0.4,
        help="Alpha in adaptive salient search objective J(k)=err(k)+alpha*err(0)*(k/n_cols) for --svd_2bit_salient_4bit_enable (default: 0.4).",
    )
    parser.add_argument(
        "--svd_2bit_salient_4bit_adaptive",
        action="store_true",
        default=False,
        help="Use adaptive salient-ratio search for --svd_2bit_salient_4bit_enable. If not set, 4bit salient columns use fixed ratio from --svd_2bit_salient_3bit_ratio.",
    )
    parser.add_argument(
        "--svd_2bit_salient_3bit_up_lim",
        type=int,
        default=50,
        help="Deprecated (kept for compatibility): previous upper bound for searched salient-column count.",
    )
    parser.add_argument(
        "--svd_one_opt",
        action="store_true",
        default=False,
        help="Enable one-step L update after R quantization: recompute L by SVD on (W - R_quantized) and quantize to 4-bit. Only works when svd_num_iters=1. Prints both old and new layer errors for comparison.",
    )
    parser.add_argument(
        "--svd_r_first",
        action="store_true",
        default=False,
        help="Enable R-first mode: quantize R first with full BiLLM, then compute L by SVD on (W - R_quantized) and quantize to 4-bit. Prints error after R quantization and final error.",
    )
    parser.add_argument(
        "--svd_r_first_refit_l",
        action="store_true",
        default=False,
        help="In --svd_r_first mode, keep first-pass R quantization fixed, then refit L by SVD on (W - R_quantized) and quantize L to 4bit (no R requantization).",
    )
    parser.add_argument(
        "--svd_early_stop",
        action="store_true",
        default=False,
        help="Enable early stopping for SVD-BiLLM iterative optimization (default: True if not specified). Stop if no improvement for 2 consecutive iterations.",
    )
    parser.add_argument(
        "--no_svd_early_stop",
        action="store_true",
        default=False,
        help="Disable early stopping for SVD-BiLLM iterative optimization (overrides --svd_early_stop).",
    )
    parser.add_argument(
        "--analyze_residual",
        action="store_true",
        default=False,
        help="Enable residual analysis (compare baseline R=W vs SVD R=W-L_quantized).",
    )
    parser.add_argument(
        "--analyze_layers",
        type=list_of_ints,
        default=[1, 14, 31],
        help="Layer indices to analyze (comma-separated, default: 1,14,31).",
    )
    parser.add_argument(
        "--analyze_save_dir",
        type=str,
        default="residual_analysis",
        help="Directory to save analysis results.",
    )
    parser.add_argument(
        "--svd_cascade_analysis",
        action="store_true",
        default=False,
        help="Run first-block cascaded SVD analysis for q/k/v/o/gate/up/down with hardcoded settings, export Excel, then exit.",
    )
    parser.add_argument(
        "--svd_residual_distribution_analysis",
        action="store_true",
        default=False,
        help="Run first-block residual distribution analysis for q/k/v/o/gate/up/down, export shared-scale 3D/PDF plots and per-row Excel stats, then exit.",
    )
    parser.add_argument(
        "--svd_residual_rank_by_vh_range",
        action="store_true",
        default=False,
        help="Deprecated. Residual analysis now only uses standard truncated SVD; this flag is ignored.",
    )
    parser.add_argument(
        "--svd_residual_fixed_rank",
        type=int,
        default=32,
        help="Residual distribution analysis only: fixed rank used in standard truncated SVD (default: 32).",
    )
    parser.add_argument(
        "--disable_salient_mask",
        action="store_true",
        default=False,
        help="Disable salient mask (structural mask), use simple block-wise binary quantization for residual R.",
    )
    parser.add_argument(
        "--row_wise_split",
        action="store_true",
        default=False,
        help="Use row-wise optimal split for non-salient weights (mask1). Each row independently searches for optimal threshold using 81 percentiles.",
    )
    parser.add_argument(
        "--column",
        action="store_true",
        default=False,
        help="Use column-wise mask split for non-salient weights. Divide mask1_2 into 4 column masks using greedy binary tree search, each mask uses 1bit quantization.",
    )
    parser.add_argument(
            '--rotation', 
            action='store_true', 
            help='是否在量化前对残差矩阵进行 Hadamard 旋转平滑'
        )
    parser.add_argument(
        "--structure_prune",
        action="store_true",
        default=False,
        help="Enable N:M structured pruning (STBLLM style) integrated with GPTQ.",
    )
    parser.add_argument(
        "--prune_n",
        type=int,
        default=0,
        help="N parameter for N:M pruning (number of weights to prune per group). Default: 0 (disabled). Recommended: 2 or 4.",
    )
    parser.add_argument(
        "--prune_m",
        type=int,
        default=0,
        help="M parameter for N:M pruning (group size). Default: 0 (disabled). Recommended: 4 or 8.",
    )
    parser.add_argument(
        "--bias_correction",
        action="store_true",
        default=False,
        help="Enable bias correction for 1-bit quantization: compensate mean shift by adding quantization error mean to layer bias.",
    )
    parser.add_argument(
        "--binary_residual",
        action="store_true",
        default=False,
        help="Use standard binary quantization for residual R (skip all masks, use full True mask + order=1 with GPTQ).",
    )
    # 正向 SmoothQuant 参数
    parser.add_argument(
        "--smoothquant_enable",
        action="store_true",
        default=False,
        help="Enable SmoothQuant: transfer activation magnitude to weights (activation quantization friendly).",
    )
    parser.add_argument(
        "--smoothquant_alpha",
        type=float,
        default=0.5,
        help="Alpha parameter for SmoothQuant scale calculation (default: 0.5).",
    )
    parser.add_argument(
        "--smoothquant_beta",
        type=float,
        default=-1,
        help="Beta parameter for SmoothQuant scale calculation (default: -1, i.e., 1-alpha).",
    )
    parser.add_argument(
        "--smoothquant_span_mode",
        type=str,
        default="absmax",
        choices=["absmax", "rms"],
        help="Span mode for SmoothQuant: 'absmax' or 'rms' (default: absmax).",
    )
    # Salient-first 模式参数
    parser.add_argument(
        "--salient_first_enable",
        action="store_true",
        default=False,
        help="Enable salient-first mode: split columns by Wanda score, quantize salient columns directly, apply SVD to non-salient columns.",
    )
    parser.add_argument(
        "--salient_first_ratio",
        type=float,
        default=0.1,
        help="Ratio of salient columns (top-k by Wanda score) in salient-first mode (default: 0.1, i.e., 10%%).",
    )
    parser.add_argument(
        "--non_salient_svd_rank",
        type=int,
        default=512,
        help="SVD rank for non-salient columns in salient-first mode (default: 512).",
    )
    parser.add_argument(
        "--salient_order",
        type=int,
        default=2,
        help="Order for salient weights (mask3) quantization (default: 2).",
    )

    args = parser.parse_args()
    import sys

    pure_gptq_requested = bool(getattr(args, "pure_gptq", False))
    if pure_gptq_requested:
        logger.warning("--pure_gptq is deprecated and now routed to --std_gptq_enable.")
    if "--block_gptq_enable" in sys.argv:
        logger.warning("--block_gptq_enable is deprecated. Use --std_gptq_enable.")
    if ("--block_gptq_static_groups" in sys.argv) or ("--block_gptq_dynamic_groups" in sys.argv):
        logger.warning(
            "--block_gptq_static_groups/--block_gptq_dynamic_groups are deprecated. "
            "Use --std_gptq_static_groups/--std_gptq_dynamic_groups."
        )

    std_group_flag_explicit = any(
        flag in sys.argv
        for flag in (
            "--std_gptq_static_groups",
            "--std_gptq_dynamic_groups",
            "--block_gptq_static_groups",
            "--block_gptq_dynamic_groups",
        )
    )
    if pure_gptq_requested and (not std_group_flag_explicit):
        # Keep legacy pure_gptq group policy behavior when no explicit std/block group flag is given.
        args.std_gptq_static_groups = bool(getattr(args, "gptq_static_groups", False))

    args.std_gptq_enable = bool(getattr(args, "std_gptq_enable", False) or pure_gptq_requested)
    # Legacy compatibility aliases used by downstream code paths.
    args.block_gptq_enable = bool(args.std_gptq_enable)
    args.block_gptq_static_groups = bool(getattr(args, "std_gptq_static_groups", True))
    # pure_gptq no longer has an independent execution path.
    args.pure_gptq = False

    if getattr(args, 'no_svd_early_stop', False):
        args.svd_early_stop = False
    elif not hasattr(args, 'svd_early_stop'):
        args.svd_early_stop = True

    svd256_then_pure_gptq2bit = bool(getattr(args, "svd256_then_pure_gptq2bit", False))
    svd256_then_pure_gptq3bit = bool(getattr(args, "svd256_then_pure_gptq3bit", False))

    if svd256_then_pure_gptq2bit and svd256_then_pure_gptq3bit:
        raise ValueError(
            "--svd256_then_pure_gptq2bit and --svd256_then_pure_gptq3bit cannot be enabled at the same time."
        )

    if svd256_then_pure_gptq2bit:
        if args.low_quant_method != "2bit":
            logger.warning("--svd256_then_pure_gptq2bit: forcing low_quant_method=2bit.")
            args.low_quant_method = "2bit"
        if not getattr(args, "std_gptq_enable", False):
            logger.warning("--svd256_then_pure_gptq2bit: forcing --std_gptq_enable=True.")
            args.std_gptq_enable = True
        if getattr(args, "disable_gptq", False):
            logger.warning("--svd256_then_pure_gptq2bit: forcing --disable_gptq=False.")
            args.disable_gptq = False
        if not getattr(args, "svd_enable", False):
            logger.warning("--svd256_then_pure_gptq2bit: forcing --svd_enable=True.")
            args.svd_enable = True
        if int(getattr(args, "svd_rank", 256)) != 256:
            logger.warning("--svd256_then_pure_gptq2bit: forcing --svd_rank=256.")
            args.svd_rank = 256
        if getattr(args, "svd_rank_by_vh_range_sigma", False):
            logger.warning("--svd256_then_pure_gptq2bit: forcing --svd_rank_by_vh_range_sigma=False.")
            args.svd_rank_by_vh_range_sigma = False
        if int(getattr(args, "svd_num_iters", 1)) != 1:
            logger.warning("--svd256_then_pure_gptq2bit: forcing --svd_num_iters=1.")
            args.svd_num_iters = 1
        if int(getattr(args, "svd_2bit_stages", 1)) != 1:
            logger.warning("--svd256_then_pure_gptq2bit: forcing --svd_2bit_stages=1.")
            args.svd_2bit_stages = 1
        if getattr(args, "svd_lowrank_fp16", False):
            logger.warning(
                "--svd256_then_pure_gptq2bit: forcing --svd_lowrank_fp16=False "
                "(low-rank branch uses 4bit per-channel)."
            )
            args.svd_lowrank_fp16 = False

    if svd256_then_pure_gptq3bit:
        if args.low_quant_method != "3bit":
            logger.warning("--svd256_then_pure_gptq3bit: forcing low_quant_method=3bit.")
            args.low_quant_method = "3bit"
        if not getattr(args, "std_gptq_enable", False):
            logger.warning("--svd256_then_pure_gptq3bit: forcing --std_gptq_enable=True.")
            args.std_gptq_enable = True
        if getattr(args, "disable_gptq", False):
            logger.warning("--svd256_then_pure_gptq3bit: forcing --disable_gptq=False.")
            args.disable_gptq = False
        if not getattr(args, "svd_enable", False):
            logger.warning("--svd256_then_pure_gptq3bit: forcing --svd_enable=True.")
            args.svd_enable = True
        if int(getattr(args, "svd_rank", 256)) != 256:
            logger.warning("--svd256_then_pure_gptq3bit: forcing --svd_rank=256.")
            args.svd_rank = 256
        if getattr(args, "svd_rank_by_vh_range_sigma", False):
            logger.warning("--svd256_then_pure_gptq3bit: forcing --svd_rank_by_vh_range_sigma=False.")
            args.svd_rank_by_vh_range_sigma = False
        if int(getattr(args, "svd_num_iters", 1)) != 1:
            logger.warning("--svd256_then_pure_gptq3bit: forcing --svd_num_iters=1.")
            args.svd_num_iters = 1
        if int(getattr(args, "svd_2bit_stages", 1)) != 1:
            logger.warning("--svd256_then_pure_gptq3bit: forcing --svd_2bit_stages=1.")
            args.svd_2bit_stages = 1
        if getattr(args, "svd_lowrank_fp16", False):
            logger.warning(
                "--svd256_then_pure_gptq3bit: forcing --svd_lowrank_fp16=False "
                "(low-rank branch uses 4bit per-channel)."
            )
            args.svd_lowrank_fp16 = False

    # Keep compatibility alias fields synchronized after preset forcing.
    args.block_gptq_enable = bool(getattr(args, "std_gptq_enable", False))
    args.block_gptq_static_groups = bool(getattr(args, "std_gptq_static_groups", True))

    if getattr(args, "svd_rank_by_vh_range_sigma", False):
        logger.warning(
            "--svd_rank_by_vh_range_sigma is deprecated and removed. "
            "Forcing --svd_rank_by_vh_range_sigma=False and using standard truncated SVD."
        )
        args.svd_rank_by_vh_range_sigma = False
    if int(getattr(args, "svd_rank_by_vh_range_sigma_topk", 32)) != 32:
        logger.warning("--svd_rank_by_vh_range_sigma_topk is deprecated and ignored.")
        args.svd_rank_by_vh_range_sigma_topk = 32
    if getattr(args, "svd_residual_rank_by_vh_range", False):
        logger.warning(
            "--svd_residual_rank_by_vh_range is deprecated and removed. "
            "Residual analysis will use standard truncated SVD."
        )
        args.svd_residual_rank_by_vh_range = False
    if getattr(args, "use_bass", False):
        logger.warning("--use_bass is deprecated and removed. Forcing --use_bass=False.")
        args.use_bass = False

    if int(getattr(args, "gptq_groupsize", -1)) == 0 or int(getattr(args, "gptq_groupsize", -1)) < -1:
        raise ValueError("--gptq_groupsize must be -1 or a positive integer.")

    if getattr(args, "gptq_true_sequential", False) and (not getattr(args, "std_gptq_enable", False)):
        logger.warning("--gptq_true_sequential is set without --std_gptq_enable; it will have no effect.")
    if getattr(args, "gptq_act_order", False) and (not getattr(args, "std_gptq_enable", False)):
        logger.warning("--gptq_act_order is set without --std_gptq_enable; it will have no effect.")

    if bool(getattr(args, "gptq_static_groups", False)):
        if not std_group_flag_explicit:
            logger.warning(
                "--gptq_static_groups is deprecated. Routing to --std_gptq_static_groups for compatibility."
            )
            args.std_gptq_static_groups = True
            args.block_gptq_static_groups = True
        else:
            logger.warning(
                "--gptq_static_groups is deprecated and ignored because explicit std/block group flags were provided."
            )

    if getattr(args, "std_gptq_enable", False):
        if args.low_quant_method not in {"2bit", "3bit", "4bit"}:
            raise ValueError("--std_gptq_enable currently only supports low_quant_method in {2bit,3bit,4bit}.")
        if getattr(args, "disable_gptq", False):
            raise ValueError("--std_gptq_enable requires GPTQ compensation. Please do not set --disable_gptq.")
        if args.low_quant_method in {"2bit", "3bit", "4bit"} and bool(getattr(args, "disable_iterative_2bit", False)):
            raise ValueError(
                "--std_gptq_enable with low_quant_method in {2bit,3bit,4bit} requires iterative affine quantization. "
                "Please remove --disable_iterative_2bit."
            )
        incompatible_flags = (
            "adaptive_rank",
            "block_greedy_rank",
            "obr",
            "disable_salient_mask",
            "row_wise_split",
            "column",
            "binary_residual",
            "smoothquant_enable",
            "salient_first_enable",
            "rotation",
            "structure_prune",
            "svd_lwc_enable",
            "svd_row_prune_offline_refill_enable",
            "gptq2bit_prune_search_enable",
            "gptq3bit_prune_search_enable",
        )
        enabled_incompatible = [flag for flag in incompatible_flags if bool(getattr(args, flag, False))]
        if enabled_incompatible:
            raise ValueError(
                "--std_gptq_enable is incompatible with: " + ", ".join(enabled_incompatible)
            )
        logger.info(
            "std_gptq mode enabled: standard column-wise GPTQ compensation "
            f"(blocksize={int(getattr(args, 'blocksize', 128))}, low_quant_method={args.low_quant_method})."
        )
        if int(getattr(args, "blocksize", 128)) != 128:
            logger.warning(
                f"std_gptq pipeline is expected to use blocksize=128, got blocksize={int(getattr(args, 'blocksize', 128))}."
            )
        else:
            logger.info("std_gptq: using fixed blocksize=128.")
        gsz = int(getattr(args, "gptq_groupsize", -1))
        if gsz == -1:
            logger.info(
                "std_gptq qparam groupsize: fallback to quantizer groupsize (default equals --blocksize)."
            )
        else:
            logger.info(f"std_gptq qparam groupsize: gptq_groupsize={gsz}.")
        logger.info(
            "std_gptq qparam policy: "
            f"{'static_groups' if bool(getattr(args, 'std_gptq_static_groups', True)) else 'dynamic_groups'}."
        )
        if args.low_quant_method in {"2bit", "3bit", "4bit"}:
            logger.info(
                f"std_gptq: iterative affine for {args.low_quant_method} is enabled "
                f"(iters={int(getattr(args, 'iterative_2bit_iters', 5))}, "
                f"fixed_iters={bool(getattr(args, 'iterative_2bit_fixed_iters', False))})."
            )
        if getattr(args, "gptq_true_sequential", False):
            logger.info(
                "std_gptq: using gptq_true_sequential grouping; "
                "qparam policy follows --std_gptq_static_groups/--std_gptq_dynamic_groups."
            )

    args.post_gptq_refill_row_ratio = min(max(float(getattr(args, "post_gptq_refill_row_ratio", 0.2)), 0.0), 1.0)
    args.post_gptq_refill_pos_ratio = min(max(float(getattr(args, "post_gptq_refill_pos_ratio", 0.2)), 0.0), 1.0)
    args.in_gptq_refill_row_ratio = min(max(float(getattr(args, "in_gptq_refill_row_ratio", 0.2)), 0.0), 1.0)
    args.in_gptq_refill_pos_ratio = min(max(float(getattr(args, "in_gptq_refill_pos_ratio", 0.2)), 0.0), 1.0)
    if bool(getattr(args, "post_gptq_refill_enable", False)) and bool(getattr(args, "in_gptq_refill_enable", False)):
        raise ValueError(
            "--post_gptq_refill_enable and --in_gptq_refill_enable are mutually exclusive. "
            "Please choose only one refill mode."
        )
    if getattr(args, "post_gptq_refill_enable", False):
        post_refill_valid = (
            bool(getattr(args, "std_gptq_enable", False))
            and bool(getattr(args, "svd_row_prune_search_enable", False))
            and str(getattr(args, "low_quant_method", "")).strip().lower() in {"2bit", "3bit", "4bit"}
        )
        if not post_refill_valid:
            logger.warning(
                "--post_gptq_refill_enable requires std_gptq + svd_row_prune_search_enable(rowprune_2pass) "
                "+ low_quant_method in {2bit,3bit,4bit}. "
                "Forcing --post_gptq_refill_enable=False."
            )
            args.post_gptq_refill_enable = False
        else:
            logger.info(
                "Post-GPTQ refill config: "
                f"row_ratio={float(args.post_gptq_refill_row_ratio):.4f}, "
                f"pos_ratio={float(args.post_gptq_refill_pos_ratio):.4f}."
            )
    if getattr(args, "in_gptq_refill_enable", False):
        in_gptq_refill_valid = (
            bool(getattr(args, "std_gptq_enable", False))
            and bool(getattr(args, "svd_row_prune_search_enable", False))
            and str(getattr(args, "low_quant_method", "")).strip().lower() in {"2bit", "3bit", "4bit"}
        )
        if not in_gptq_refill_valid:
            logger.warning(
                "--in_gptq_refill_enable requires std_gptq + svd_row_prune_search_enable(rowprune_2pass) "
                "+ low_quant_method in {2bit,3bit,4bit}. "
                "Forcing --in_gptq_refill_enable=False."
            )
            args.in_gptq_refill_enable = False
        else:
            logger.info(
                "In-GPTQ refill config: "
                f"row_ratio={float(args.in_gptq_refill_row_ratio):.4f}, "
                f"pos_ratio={float(args.in_gptq_refill_pos_ratio):.4f}."
            )

    if getattr(args, "gptq3bit_prune_search_enable", False):
        raise ValueError(
            "--gptq3bit_prune_search_enable is deprecated and removed. "
            "Use --svd_row_prune_search_enable --svd_row_prune_quant_scheme std3bit with --std_gptq_enable."
        )

    if getattr(args, "gptq2bit_prune_search_enable", False):
        raise ValueError(
            "--gptq2bit_prune_search_enable is deprecated and removed. "
            "Use --svd_row_prune_search_enable --svd_row_prune_quant_scheme std2bit with --std_gptq_enable."
        )

    if getattr(args, "svd_cascade_analysis", False) and getattr(args, "svd_residual_distribution_analysis", False):
        raise ValueError("--svd_cascade_analysis and --svd_residual_distribution_analysis cannot be enabled together.")

    if getattr(args, "svd_residual_distribution_analysis", False):
        if not getattr(args, "svd_enable", False):
            logger.warning("--svd_residual_distribution_analysis requires --svd_enable. Forcing --svd_enable=True.")
            args.svd_enable = True
        if int(getattr(args, "svd_residual_fixed_rank", 32)) <= 0:
            raise ValueError("--svd_residual_fixed_rank must be > 0.")

    if getattr(args, "svd_lwc_enable", False):
        if args.low_quant_method != "2bit":
            raise ValueError("--svd_lwc_enable currently only supports low_quant_method=2bit.")
        if getattr(args, "obr", False):
            raise ValueError("--svd_lwc_enable currently targets GPTQ residual path and is incompatible with --obr.")
        if not getattr(args, "svd_enable", False):
            logger.warning("--svd_lwc_enable requires --svd_enable. Forcing --svd_enable=True.")
            args.svd_enable = True

        candidates = [float(x) for x in getattr(args, "svd_lwc_candidates", [1.0, 0.999, 0.995, 0.99, 0.98])]
        if len(candidates) == 0:
            candidates = [1.0, 0.999, 0.995, 0.99, 0.98]
        candidates = [min(max(float(x), 0.0), 1.0) for x in candidates]
        if all(abs(x - 1.0) > 1e-12 for x in candidates):
            candidates = [1.0] + candidates
        args.svd_lwc_candidates = candidates
        logger.info(
            "svd_lwc_enable enabled: per-block clip-ratio search before standard 2bit per-channel residual quantization, "
            f"candidates={args.svd_lwc_candidates}."
        )

    if getattr(args, "svd_row_prune_search_enable", False):
        low_method = str(getattr(args, "low_quant_method", "")).strip().lower()
        if low_method not in {"2bit", "3bit", "4bit"}:
            raise ValueError("--svd_row_prune_search_enable currently only supports low_quant_method in {2bit,3bit,4bit}.")
        if low_method in {"3bit", "4bit"} and (not bool(getattr(args, "std_gptq_enable", False))):
            raise ValueError(
                f"--svd_row_prune_search_enable with low_quant_method={low_method} currently requires --std_gptq_enable."
            )
        if getattr(args, "obr", False):
            raise ValueError("--svd_row_prune_search_enable currently targets GPTQ residual path and is incompatible with --obr.")
        if not getattr(args, "svd_enable", False):
            if bool(getattr(args, "std_gptq_enable", False)):
                logger.info(
                    "--svd_row_prune_search_enable with --std_gptq_enable and --svd_enable=False: "
                    "running no-SVD block residual path (R=W), i.e. row-prune search + std_gptq on full weight."
                )
            else:
                logger.warning("--svd_row_prune_search_enable requires --svd_enable. Forcing --svd_enable=True.")
                args.svd_enable = True

        candidates = [int(x) for x in getattr(args, "svd_row_prune_m_candidates", [0, 4, 8, 16, 32, 64])]
        if len(candidates) == 0:
            candidates = [0, 4, 8, 16, 32, 64]
        candidates = [max(int(x), 0) for x in candidates]
        dedup_candidates = []
        for m in candidates:
            if m not in dedup_candidates:
                dedup_candidates.append(m)
        args.svd_row_prune_m_candidates = dedup_candidates
        row_prune_score_metric = str(getattr(args, "svd_row_prune_score_metric", "magnitude")).strip().lower()
        if row_prune_score_metric not in {"magnitude", "wanda"}:
            raise ValueError("--svd_row_prune_score_metric must be one of: magnitude/wanda.")
        args.svd_row_prune_score_metric = row_prune_score_metric

        if getattr(args, "svd_lwc_enable", False):
            logger.warning("--svd_row_prune_search_enable and --svd_lwc_enable are both set. Disabling --svd_lwc_enable.")
            args.svd_lwc_enable = False
        row_prune_order = int(getattr(args, "svd_row_prune_binary_order", 2))
        if row_prune_order not in (1, 2, 3):
            raise ValueError("--svd_row_prune_binary_order must be 1, 2, or 3.")
        args.svd_row_prune_binary_order = row_prune_order
        row_prune_quant_scheme = str(getattr(args, "svd_row_prune_quant_scheme", "binary")).strip().lower()
        if row_prune_quant_scheme not in {"binary", "std2bit", "std3bit", "std4bit"}:
            raise ValueError("--svd_row_prune_quant_scheme must be one of: binary/std2bit/std3bit/std4bit.")
        args.svd_row_prune_quant_scheme = row_prune_quant_scheme
        row_prune_clip_enable = bool(getattr(args, "svd_row_prune_clip_search_enable", False))
        args.svd_row_prune_clip_search_enable = row_prune_clip_enable
        clip_candidates = [
            float(x) for x in getattr(args, "svd_row_prune_clip_candidates", [1.0, 0.999, 0.995, 0.99, 0.98, 0.96])
        ]
        if len(clip_candidates) == 0:
            clip_candidates = [1.0, 0.999, 0.995, 0.99, 0.98, 0.96]
        clip_candidates = [min(max(float(x), 0.0), 1.0) for x in clip_candidates]
        dedup_clip_candidates = []
        for p in clip_candidates:
            if all(abs(p - x) > 1e-12 for x in dedup_clip_candidates):
                dedup_clip_candidates.append(p)
        if all(abs(x - 1.0) > 1e-12 for x in dedup_clip_candidates):
            dedup_clip_candidates = [1.0] + dedup_clip_candidates
        args.svd_row_prune_clip_candidates = dedup_clip_candidates
        clip_min_value = float(getattr(args, "svd_row_prune_clip_min_value", 1e-8))
        if clip_min_value < 0:
            raise ValueError("--svd_row_prune_clip_min_value must be >= 0.")
        args.svd_row_prune_clip_min_value = clip_min_value
        act_topk = int(getattr(args, "svd_row_prune_act_topk", 3))
        if act_topk < 0:
            raise ValueError("--svd_row_prune_act_topk must be >= 0.")
        if row_prune_quant_scheme == "binary" and act_topk > 0:
            logger.warning(
                "--svd_row_prune_act_topk is only supported with --svd_row_prune_quant_scheme=std2bit/std3bit/std4bit. "
                "Current scheme is binary, forcing --svd_row_prune_act_topk=0."
            )
            act_topk = 0
        if row_prune_clip_enable and act_topk > 0:
            logger.warning(
                "--svd_row_prune_clip_search_enable conflicts with --svd_row_prune_act_topk>0; "
                "forcing --svd_row_prune_clip_search_enable=False."
            )
            row_prune_clip_enable = False
            args.svd_row_prune_clip_search_enable = False
        if bool(getattr(args, "std_gptq_enable", False)) and act_topk > 0:
            logger.warning(
                "std_gptq rowprune_2pass now disables legacy act_topk one-shot refill rerank. "
                "Forcing --svd_row_prune_act_topk=0."
            )
            act_topk = 0
        args.svd_row_prune_act_topk = act_topk
        refill_enable = bool(getattr(args, "svd_row_prune_offline_refill_enable", False))
        args.svd_row_prune_offline_refill_enable = refill_enable
        refill_sweeps = int(getattr(args, "svd_row_prune_offline_refill_sweeps", 1))
        if refill_sweeps < 1:
            raise ValueError("--svd_row_prune_offline_refill_sweeps must be >= 1.")
        args.svd_row_prune_offline_refill_sweeps = refill_sweeps
        refill_max_positions = int(getattr(args, "svd_row_prune_offline_refill_max_positions", -1))
        args.svd_row_prune_offline_refill_max_positions = refill_max_positions
        args.svd_row_prune_offline_refill_include_zero = bool(
            getattr(args, "svd_row_prune_offline_refill_include_zero", True)
        )
        if row_prune_quant_scheme != "binary":
            logger.info(
                "--svd_row_prune_binary_order is ignored when --svd_row_prune_quant_scheme is std2bit/std3bit/std4bit."
            )
        if row_prune_quant_scheme == "binary":
            row_prune_quant_desc = f"binary residual approximation (order={row_prune_order})"
        elif row_prune_quant_scheme == "std2bit":
            row_prune_quant_desc = "standard 2bit per-channel quantization"
        elif row_prune_quant_scheme == "std3bit":
            row_prune_quant_desc = "standard 3bit per-channel quantization"
        else:
            row_prune_quant_desc = "standard 4bit per-channel quantization"
        logger.info(
            "svd_row_prune_search_enable enabled: per-block row-prune m-search, "
            f"candidates={args.svd_row_prune_m_candidates}, "
            f"score_metric={args.svd_row_prune_score_metric}, "
            "objective=L(m)=sum((R_b-Q_b)^2) with original unpruned R_b, "
            f"Q_b via {row_prune_quant_desc}."
        )
        if int(args.svd_row_prune_act_topk) > 0:
            logger.info(
                "svd_row_prune_act_topk enabled: top-K candidate rerank with one-shot RTN refill "
                "on pruned positions, selection objective S_act=sum((E@H_b)*E), "
                f"topk={int(args.svd_row_prune_act_topk)}."
            )
        if row_prune_clip_enable:
            logger.info(
                "svd_row_prune_clip_search_enable enabled: second-stage clip search after best m selection, "
                f"clip_candidates={args.svd_row_prune_clip_candidates}, "
                f"clip_min_value={float(args.svd_row_prune_clip_min_value):.3e}."
            )
        if refill_enable:
            refill_pos_desc = (
                "all pruned positions"
                if int(args.svd_row_prune_offline_refill_max_positions) <= 0
                else f"top-{int(args.svd_row_prune_offline_refill_max_positions)} pruned positions"
            )
            logger.info(
                "svd_row_prune_offline_refill_enable enabled: offline full-fill refill with full-row activation objective "
                "DeltaL=-2*Delta*g_j+Delta^2*H_jj, "
                f"sweeps={int(args.svd_row_prune_offline_refill_sweeps)}, "
                f"positions={refill_pos_desc}, "
                f"include_zero={bool(args.svd_row_prune_offline_refill_include_zero)}."
            )
    elif getattr(args, "svd_row_prune_clip_search_enable", False):
        logger.warning(
            "--svd_row_prune_clip_search_enable requires --svd_row_prune_search_enable. "
            "Forcing --svd_row_prune_clip_search_enable=False."
        )
        args.svd_row_prune_clip_search_enable = False

    if (not getattr(args, "svd_row_prune_search_enable", False)) and getattr(
        args, "svd_row_prune_offline_refill_enable", False
    ):
        logger.warning(
            "--svd_row_prune_offline_refill_enable requires --svd_row_prune_search_enable. "
            "Forcing --svd_row_prune_offline_refill_enable=False."
        )
        args.svd_row_prune_offline_refill_enable = False

    if getattr(args, "svd_auto_rank_deltae_stop", False):
        if args.low_quant_method != "2bit":
            raise ValueError("--svd_auto_rank_deltae_stop currently only supports low_quant_method=2bit.")
        if not getattr(args, "svd_enable", False):
            logger.warning("--svd_auto_rank_deltae_stop requires --svd_enable. Forcing --svd_enable=True.")
            args.svd_enable = True

        # Hardcoded runtime behavior requested by user:
        # - single-stage SVD
        # - rank search starts from 16 with step 16, stop if Delta(normalized act error) < 0.05
        # - residual quantization uses full-block order=2 binary approximation only
        if int(getattr(args, "svd_num_iters", 1)) != 1:
            logger.warning("--svd_auto_rank_deltae_stop: forcing --svd_num_iters=1.")
            args.svd_num_iters = 1
        if int(getattr(args, "svd_2bit_stages", 1)) != 1:
            logger.warning("--svd_auto_rank_deltae_stop: forcing --svd_2bit_stages=1.")
            args.svd_2bit_stages = 1
        if int(getattr(args, "svd_rank", 16)) != 16:
            logger.warning("--svd_auto_rank_deltae_stop: forcing --svd_rank=16 as search start anchor.")
            args.svd_rank = 16

        forced_false_flags = (
            "adaptive_rank",
            "block_greedy_rank",
            "use_bass",
            "svd_binary_refit_enable",
            "svd_2bit_extra_8bit_stage_enable",
            "svd_2bit_salient_3bit_enable",
            "svd_2bit_salient_4bit_enable",
            "svd_2bit_obr_cascade_enable",
            "svd_2bit_obr_twogroup_enable",
            "svd_2bit_obr_twogroup_adaptive_enable",
            "svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio",
            "svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor",
            "svd_2bit_obr_4group_enable",
            "obr",
            "binary_residual",
            "disable_salient_mask",
            "row_wise_split",
            "column",
            "svd_obr_cascade_lowrank_8bit_enable",
            "svd_obr_interblock_gptq_propagation_disable",
        )
        for flag in forced_false_flags:
            if getattr(args, flag, False):
                logger.warning(f"--svd_auto_rank_deltae_stop: forcing --{flag}=False.")
                setattr(args, flag, False)

        logger.info(
            "svd_auto_rank_deltae_stop enabled: single-stage SVD rank auto-search "
            "(start=16, step=16, stop if DeltaE_norm < 0.05) + full-block order=2 residual path."
        )

    if getattr(args, 'svd_energy_stats_enable', False):
        if getattr(args, 'svd_energy_probe_rank', 128) <= 0:
            raise ValueError("--svd_energy_probe_rank must be > 0.")
        thresholds = getattr(args, 'svd_energy_thresholds', [0.90, 0.92])
        if thresholds is None or len(thresholds) == 0:
            raise ValueError("--svd_energy_thresholds cannot be empty when --svd_energy_stats_enable is set.")
        for tau in thresholds:
            tau_val = float(tau)
            if tau_val <= 0.0 or tau_val > 1.0:
                raise ValueError("Each value in --svd_energy_thresholds must be in (0, 1].")
        if not getattr(args, 'svd_enable', False):
            logger.warning(
                "--svd_energy_stats_enable is set while --svd_enable is disabled. "
                "No SVD energy stats will be collected."
            )

    if getattr(args, 'svd_bit_stats_enable', False):
        if float(getattr(args, 'svd_bit_stats_scale_bits', 16)) < 0:
            raise ValueError("--svd_bit_stats_scale_bits must be >= 0.")
        if float(getattr(args, 'svd_bit_stats_zp_bits', 8)) < 0:
            raise ValueError("--svd_bit_stats_zp_bits must be >= 0.")
        if float(getattr(args, 'svd_bit_stats_binary_data_bits', 2.0)) < 0:
            raise ValueError("--svd_bit_stats_binary_data_bits must be >= 0.")
        if not getattr(args, 'svd_enable', False):
            logger.warning(
                "--svd_bit_stats_enable is set while --svd_enable is disabled. "
                "No SVD bit stats will be collected."
            )

    if getattr(args, 'obr', False):
        if args.low_quant_method != "2bit":
            raise ValueError("--obr currently only supports low_quant_method=2bit.")
        if not getattr(args, 'svd_enable', False):
            logger.warning(
                "--obr is running without --svd_enable. "
                "This enables no-SVD OBR residual-only quantization for comparison."
            )
        if not getattr(args, 'disable_gptq', False):
            raise ValueError("--obr is mutually exclusive with GPTQ. Please set --disable_gptq.")
        if getattr(args, 'svd_2bit_salient_3bit_enable', False):
            raise ValueError(
                "--obr does not support --svd_2bit_salient_3bit_enable directly. "
                "Use --svd_2bit_salient_4bit_enable or --svd_2bit_obr_cascade_enable."
            )

    svd_obr_scheme = str(getattr(args, 'svd_obr_scheme', 'none')).strip().lower()
    if svd_obr_scheme not in {'none', 'adp3_fix01', 'search4_linked', 'adp4_linked'}:
        raise ValueError("--svd_obr_scheme must be one of: none/adp3_fix01/search4_linked/adp4_linked.")

    if svd_obr_scheme != 'none':
        if args.low_quant_method != '2bit':
            raise ValueError(f"--svd_obr_scheme {svd_obr_scheme} currently only supports low_quant_method=2bit.")
        if not getattr(args, 'svd_enable', False):
            raise ValueError(f"--svd_obr_scheme {svd_obr_scheme} requires --svd_enable.")
        if not getattr(args, 'obr', False):
            raise ValueError(f"--svd_obr_scheme {svd_obr_scheme} requires --obr.")
        if not getattr(args, 'disable_gptq', False):
            raise ValueError(f"--svd_obr_scheme {svd_obr_scheme} requires --disable_gptq.")

        args.svd_2bit_obr_twogroup_enable = True
        args.svd_2bit_salient_3bit_enable = False
        args.svd_2bit_salient_4bit_enable = False
        args.svd_2bit_obr_cascade_enable = False
        args.svd_2bit_obr_4group_enable = False
        args.svd_2bit_obr_twogroup_adaptive_enable = False
        args.svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio = False
        args.svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor = False

        if not getattr(args, 'svd_obr_interblock_gptq_propagation_disable', False):
            logger.info(f"--svd_obr_scheme={svd_obr_scheme}: forcing --svd_obr_interblock_gptq_propagation_disable for fair comparison.")
        args.svd_obr_interblock_gptq_propagation_disable = True

        if svd_obr_scheme == 'adp3_fix01':
            args.svd_2bit_obr_twogroup_adaptive_enable = True
            args.svd_2bit_obr_twogroup_adaptive_max_stages = 3
            args.svd_2bit_obr_twogroup_adaptive_base_ratio = 0.1
            args.svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio = True
            args.svd_2bit_obr_twogroup_4bit_ratio = 0.1
        elif svd_obr_scheme == 'adp4_linked':
            args.svd_2bit_obr_twogroup_adaptive_enable = True
            args.svd_2bit_obr_twogroup_adaptive_max_stages = 4
            args.svd_2bit_obr_twogroup_adaptive_base_ratio = 0.1
            args.svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor = True
        elif svd_obr_scheme == 'search4_linked':
            args.svd_2bit_obr_twogroup_4bit_ratio = 0.1

        logger.info(
            f"Enabled OBR scheme preset: {svd_obr_scheme} "
            f"(twogroup={getattr(args, 'svd_2bit_obr_twogroup_enable', False)}, "
            f"adaptive={getattr(args, 'svd_2bit_obr_twogroup_adaptive_enable', False)}, "
            f"max_stages={getattr(args, 'svd_2bit_obr_twogroup_adaptive_max_stages', 3)}, "
            f"ratio4={getattr(args, 'svd_2bit_obr_twogroup_4bit_ratio', 0.1):.2f})."
        )

    if getattr(args, 'svd_2bit_obr_cascade_enable', False):
        if args.low_quant_method != "2bit":
            raise ValueError("--svd_2bit_obr_cascade_enable currently only supports low_quant_method=2bit.")
        if not getattr(args, 'svd_enable', False):
            raise ValueError("--svd_2bit_obr_cascade_enable requires --svd_enable.")
        if not getattr(args, 'obr', False):
            raise ValueError("--svd_2bit_obr_cascade_enable requires --obr.")
        if not getattr(args, 'disable_gptq', False):
            raise ValueError("--svd_2bit_obr_cascade_enable requires --disable_gptq.")
        if getattr(args, 'svd_2bit_salient_3bit_enable', False) or getattr(args, 'svd_2bit_salient_4bit_enable', False):
            raise ValueError("--svd_2bit_obr_cascade_enable cannot be used with --svd_2bit_salient_3bit_enable or --svd_2bit_salient_4bit_enable.")
        ratio4 = float(getattr(args, 'svd_2bit_obr_cascade_4bit_ratio', 0.1))
        ratio3 = float(getattr(args, 'svd_2bit_obr_cascade_3bit_ratio', 0.1))
        if ratio4 < 0.0 or ratio4 > 1.0 or ratio3 < 0.0 or ratio3 > 1.0:
            raise ValueError("--svd_2bit_obr_cascade_4bit_ratio and --svd_2bit_obr_cascade_3bit_ratio must be in [0,1].")
        if ratio4 + ratio3 > 1.0 + 1e-8:
            raise ValueError("--svd_2bit_obr_cascade_4bit_ratio + --svd_2bit_obr_cascade_3bit_ratio must be <= 1.")

    if getattr(args, 'svd_2bit_obr_4group_enable', False):
        if args.low_quant_method != "2bit":
            raise ValueError("--svd_2bit_obr_4group_enable currently only supports low_quant_method=2bit.")
        if not getattr(args, 'svd_enable', False):
            raise ValueError("--svd_2bit_obr_4group_enable requires --svd_enable.")
        if not getattr(args, 'obr', False):
            raise ValueError("--svd_2bit_obr_4group_enable requires --obr.")
        if not getattr(args, 'disable_gptq', False):
            raise ValueError("--svd_2bit_obr_4group_enable requires --disable_gptq.")
        if (
            getattr(args, 'svd_2bit_salient_3bit_enable', False)
            or getattr(args, 'svd_2bit_salient_4bit_enable', False)
            or getattr(args, 'svd_2bit_obr_cascade_enable', False)
        ):
            raise ValueError(
                "--svd_2bit_obr_4group_enable cannot be used with "
                "--svd_2bit_salient_3bit_enable, --svd_2bit_salient_4bit_enable, "
                "or --svd_2bit_obr_cascade_enable."
            )
        ratio4 = float(getattr(args, 'svd_2bit_obr_4group_4bit_ratio', 0.1))
        ratio2_1 = float(getattr(args, 'svd_2bit_obr_4group_2bit_ratio_1', 0.3))
        ratio2_2 = float(getattr(args, 'svd_2bit_obr_4group_2bit_ratio_2', 0.3))
        if (
            ratio4 < 0.0 or ratio4 > 1.0
            or ratio2_1 < 0.0 or ratio2_1 > 1.0
            or ratio2_2 < 0.0 or ratio2_2 > 1.0
        ):
            raise ValueError(
                "--svd_2bit_obr_4group_4bit_ratio, --svd_2bit_obr_4group_2bit_ratio_1, "
                "and --svd_2bit_obr_4group_2bit_ratio_2 must be in [0,1]."
            )
        if ratio4 + ratio2_1 + ratio2_2 > 1.0 + 1e-8:
            raise ValueError(
                "--svd_2bit_obr_4group_4bit_ratio + --svd_2bit_obr_4group_2bit_ratio_1 + "
                "--svd_2bit_obr_4group_2bit_ratio_2 must be <= 1."
            )
        scheme = str(getattr(args, 'svd_2bit_obr_4group_non4bit_scheme', 'order2')).strip().lower()
        if scheme not in {"order2", "std2bit", "ternary", "ttb1"}:
            raise ValueError(
                "--svd_2bit_obr_4group_non4bit_scheme must be one of "
                "order2/std2bit/ternary/ttb1."
            )
        delta_factor = float(getattr(args, 'svd_2bit_obr_ternary_delta_factor', 0.7))
        if delta_factor < 0.0:
            raise ValueError("--svd_2bit_obr_ternary_delta_factor must be >= 0.")

    if getattr(args, 'svd_2bit_obr_twogroup_adaptive_enable', False):
        args.svd_2bit_obr_twogroup_enable = True
        if args.low_quant_method != "2bit":
            raise ValueError("--svd_2bit_obr_twogroup_adaptive_enable currently only supports low_quant_method=2bit.")
        if not getattr(args, 'svd_enable', False):
            raise ValueError("--svd_2bit_obr_twogroup_adaptive_enable requires --svd_enable.")
        if not getattr(args, 'obr', False):
            raise ValueError("--svd_2bit_obr_twogroup_adaptive_enable requires --obr.")
        if not getattr(args, 'disable_gptq', False):
            raise ValueError("--svd_2bit_obr_twogroup_adaptive_enable requires --disable_gptq.")
        if (
            getattr(args, 'svd_2bit_salient_3bit_enable', False)
            or getattr(args, 'svd_2bit_salient_4bit_enable', False)
            or getattr(args, 'svd_2bit_obr_cascade_enable', False)
            or getattr(args, 'svd_2bit_obr_4group_enable', False)
        ):
            raise ValueError(
                "--svd_2bit_obr_twogroup_adaptive_enable cannot be used with "
                "--svd_2bit_salient_3bit_enable, --svd_2bit_salient_4bit_enable, "
                "--svd_2bit_obr_cascade_enable, or --svd_2bit_obr_4group_enable."
            )
        max_stages = int(getattr(args, 'svd_2bit_obr_twogroup_adaptive_max_stages', 3))
        if max_stages < 1 or max_stages > 4:
            raise ValueError("--svd_2bit_obr_twogroup_adaptive_max_stages must be in [1,4].")
        e128_threshold = float(getattr(args, 'svd_2bit_obr_twogroup_adaptive_e128_threshold', 0.2))
        if e128_threshold < 0.0 or e128_threshold > 1.0:
            raise ValueError("--svd_2bit_obr_twogroup_adaptive_e128_threshold must be in [0,1].")
        base_ratio = float(getattr(args, 'svd_2bit_obr_twogroup_adaptive_base_ratio', 0.1))
        if base_ratio < 0.0:
            raise ValueError("--svd_2bit_obr_twogroup_adaptive_base_ratio must be >= 0.")
        if bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio', False)):
            ratio4 = float(getattr(args, 'svd_2bit_obr_twogroup_4bit_ratio', 0.1))
            if ratio4 < 0.0 or ratio4 > 1.0:
                raise ValueError("--svd_2bit_obr_twogroup_4bit_ratio must be in [0,1] when --svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio is set.")
        if bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor', False)) and bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio', False)):
            raise ValueError("--svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor cannot be used together with --svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio.")

    if bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio', False)) and not bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_enable', False)):
        logger.warning("--svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio is set without --svd_2bit_obr_twogroup_adaptive_enable; it will have no effect.")

    if bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor', False)) and not bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_enable', False)):
        logger.warning("--svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor is set without --svd_2bit_obr_twogroup_adaptive_enable; it will have no effect.")

    if getattr(args, 'svd_2bit_obr_twogroup_enable', False):
        if args.low_quant_method != "2bit":
            raise ValueError("--svd_2bit_obr_twogroup_enable currently only supports low_quant_method=2bit.")
        if not getattr(args, 'svd_enable', False):
            raise ValueError("--svd_2bit_obr_twogroup_enable requires --svd_enable.")
        if not getattr(args, 'obr', False):
            raise ValueError("--svd_2bit_obr_twogroup_enable requires --obr.")
        if not getattr(args, 'disable_gptq', False):
            raise ValueError("--svd_2bit_obr_twogroup_enable requires --disable_gptq.")
        if (
            getattr(args, 'svd_2bit_salient_3bit_enable', False)
            or getattr(args, 'svd_2bit_salient_4bit_enable', False)
            or getattr(args, 'svd_2bit_obr_cascade_enable', False)
            or getattr(args, 'svd_2bit_obr_4group_enable', False)
        ):
            raise ValueError(
                "--svd_2bit_obr_twogroup_enable cannot be used with "
                "--svd_2bit_salient_3bit_enable, --svd_2bit_salient_4bit_enable, "
                "--svd_2bit_obr_cascade_enable, or --svd_2bit_obr_4group_enable."
            )
        ratio4 = float(getattr(args, 'svd_2bit_obr_twogroup_4bit_ratio', 0.1))
        if ratio4 < 0.0 or ratio4 > 1.0:
            raise ValueError("--svd_2bit_obr_twogroup_4bit_ratio must be in [0,1].")

    if getattr(args, 'svd_obr_cascade_lowrank_8bit_enable', False):
        if args.low_quant_method != "2bit":
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable currently only supports low_quant_method=2bit.")
        if not getattr(args, 'svd_enable', False):
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable requires --svd_enable.")
        if not getattr(args, 'svd_2bit_obr_cascade_enable', False):
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable requires --svd_2bit_obr_cascade_enable.")
        if not getattr(args, 'obr', False):
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable requires --obr.")
        if not getattr(args, 'disable_gptq', False):
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable requires --disable_gptq.")
        if getattr(args, 'svd_2bit_stages', 2) != 1:
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable requires --svd_2bit_stages 1.")
        if getattr(args, 'svd_num_iters', 1) != 1:
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable requires --svd_num_iters 1.")
        if getattr(args, 'svd_2bit_extra_8bit_stage_enable', False):
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable is incompatible with --svd_2bit_extra_8bit_stage_enable.")
        if getattr(args, 'svd_drop_residual', False):
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable is incompatible with --svd_drop_residual.")
        if getattr(args, 'svd_r_first', False):
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable is incompatible with --svd_r_first.")
        if getattr(args, 'svd_one_opt', False):
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable is incompatible with --svd_one_opt.")
        if getattr(args, 'rotation', False):
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable is incompatible with --rotation.")
        if getattr(args, 'svd_2bit_obr_4group_enable', False):
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable is incompatible with --svd_2bit_obr_4group_enable.")
        if getattr(args, 'svd_2bit_obr_twogroup_enable', False):
            raise ValueError("--svd_obr_cascade_lowrank_8bit_enable is incompatible with --svd_2bit_obr_twogroup_enable.")
        gamma = float(getattr(args, 'svd_obr_cascade_lowrank_gamma', 1.0))
        damp = float(getattr(args, 'svd_obr_cascade_lowrank_damp', 0.01))
        if gamma < 0.0:
            raise ValueError("--svd_obr_cascade_lowrank_gamma must be >= 0.")
        if damp < 0.0:
            raise ValueError("--svd_obr_cascade_lowrank_damp must be >= 0.")

    if getattr(args, 'svd_2bit_extra_8bit_stage_enable', False):
        if args.low_quant_method != "2bit":
            raise ValueError("--svd_2bit_extra_8bit_stage_enable currently only supports low_quant_method=2bit.")
        if not getattr(args, 'svd_enable', False):
            raise ValueError("--svd_2bit_extra_8bit_stage_enable requires --svd_enable.")
        if getattr(args, 'svd_2bit_extra_8bit_stage_rank', 64) <= 0:
            raise ValueError("--svd_2bit_extra_8bit_stage_rank must be > 0.")

    if getattr(args, 'svd_binary_refit_enable', False):
        if not getattr(args, 'svd_enable', False):
            raise ValueError("--svd_binary_refit_enable requires --svd_enable.")
        if getattr(args, 'obr', False):
            raise ValueError("--svd_binary_refit_enable is incompatible with --obr.")
        if getattr(args, 'svd_2bit_obr_cascade_enable', False):
            raise ValueError("--svd_binary_refit_enable is incompatible with --svd_2bit_obr_cascade_enable.")
        if getattr(args, 'svd_2bit_extra_8bit_stage_enable', False):
            raise ValueError("--svd_binary_refit_enable is incompatible with --svd_2bit_extra_8bit_stage_enable.")
        if getattr(args, 'svd_2bit_salient_3bit_enable', False) or getattr(args, 'svd_2bit_salient_4bit_enable', False):
            raise ValueError("--svd_binary_refit_enable uses no-mask binary residual. Disable salient split flags.")
        if getattr(args, 'svd_drop_residual', False):
            raise ValueError("--svd_binary_refit_enable is incompatible with --svd_drop_residual.")
        if getattr(args, 'svd_r_first', False):
            raise ValueError("--svd_binary_refit_enable is incompatible with --svd_r_first.")
        if getattr(args, 'svd_one_opt', False):
            raise ValueError("--svd_binary_refit_enable is incompatible with --svd_one_opt.")
        if getattr(args, 'rotation', False):
            raise ValueError("--svd_binary_refit_enable is incompatible with --rotation.")
        if getattr(args, 'structure_prune', False):
            raise ValueError("--svd_binary_refit_enable is incompatible with --structure_prune.")

    if getattr(args, 'svd_r_first_refit_l', False):
        if not getattr(args, 'svd_r_first', False):
            logger.warning("--svd_r_first_refit_l requires --svd_r_first. Forcing --svd_r_first_refit_l=False.")
            args.svd_r_first_refit_l = False
        if not getattr(args, 'svd_enable', False):
            logger.warning("--svd_r_first_refit_l requires --svd_enable. Forcing --svd_enable=True.")
            args.svd_enable = True
    
    groupsize = args.blocksize

    device = args.device
    
    model_name = args.model.split('/')[-1] if '/' in args.model else args.model
    
    parts = [model_name]
    
    parts.append(args.low_quant_method)

    if getattr(args, 'std_gptq_enable', False):
        parts.append("stdgptq")
        gsz = int(getattr(args, 'gptq_groupsize', -1))
        if gsz != -1:
            parts.append(f"g{gsz}")
        if bool(getattr(args, 'gptq_act_order', False)):
            parts.append("actord")
        if bool(getattr(args, 'gptq_true_sequential', False)):
            parts.append("tseq")
        if bool(getattr(args, 'gptq_sym', False)):
            parts.append("sym")
        if bool(getattr(args, 'gptq2bit_prune_search_enable', False)) or bool(getattr(args, 'gptq3bit_prune_search_enable', False)):
            parts.append("prunesrch")
    
    parts.append(f"b{groupsize}")
    
    if args.salient_metric != "magnitude":  # 假设 magnitude 是默认值
        parts.append(args.salient_metric[:3])  # 缩写：mag, hes
    
    if bool(getattr(args, 'svd_auto_rank_deltae_stop', False)):
        parts.append('svdauto')

    if getattr(args, 'svd_enable', False):
        if getattr(args, 'adaptive_rank', False):
            parts.append(f"adaptive")
            parts.append(f"avg{getattr(args, 'avg_rank', 256)}")
            
            # ⭐ Add block_greedy_rank information
            if getattr(args, 'block_greedy_rank', False):
                # Check if two-level allocation is enabled
                disable_two_level = getattr(args, 'disable_two_level_allocation', False)
                if disable_two_level:
                    # Single-level allocation
                    parts.append("blk_greedy_1L")
                    parts.append(f"{getattr(args, 'block_greedy_base_rank', 128)}-{getattr(args, 'block_greedy_max_rank', 512)}")
                else:
                    # Two-level allocation
                    parts.append("blk_greedy_2L")
                    # Add block-level range (min_rank-max_rank)
                    parts.append(f"L1_{getattr(args, 'min_rank', 128)}-{getattr(args, 'max_rank', 512)}")
                    # Add layer-level range (base_rank-max_rank)
                    parts.append(f"L2_{getattr(args, 'block_greedy_base_rank', 128)}-{getattr(args, 'block_greedy_max_rank', 512)}")
                
                # Add stable rank scoring if enabled
                if getattr(args, 'use_activation_entropy', False):
                    parts.append("stable_rank")
        else:
            parts.append(f"svd{getattr(args, 'svd_rank', 8)}")
        # Add iterative optimization information
        svd_num_iters = getattr(args, 'svd_num_iters', 1)
        if svd_num_iters > 1:
            parts.append(f"iter{svd_num_iters}")
        if getattr(args, 'svd_drop_residual', False):
            parts.append("dropR")
        if getattr(args, 'obr', False):
            parts.append("obr")
        if getattr(args, 'disable_gptq', False):
            parts.append("nogptq")
        else:
            parts.append("gptq")
        svd_obr_scheme = str(getattr(args, 'svd_obr_scheme', 'none')).strip().lower()
        if svd_obr_scheme != 'none':
            parts.append(f"sch_{svd_obr_scheme}")
        if getattr(args, 'svd_2bit_obr_cascade_enable', False):
            parts.append(
                f"casc43_r4{getattr(args, 'svd_2bit_obr_cascade_4bit_ratio', 0.1):.2f}"
                f"_r3{getattr(args, 'svd_2bit_obr_cascade_3bit_ratio', 0.1):.2f}"
            )
        if getattr(args, 'svd_2bit_obr_4group_enable', False):
            parts.append(
                f"casc4222_r4{getattr(args, 'svd_2bit_obr_4group_4bit_ratio', 0.1):.2f}"
                f"_r2a{getattr(args, 'svd_2bit_obr_4group_2bit_ratio_1', 0.3):.2f}"
                f"_r2b{getattr(args, 'svd_2bit_obr_4group_2bit_ratio_2', 0.3):.2f}"
            )
            parts.append(f"n4_{getattr(args, 'svd_2bit_obr_4group_non4bit_scheme', 'order2')}")
            if str(getattr(args, 'svd_2bit_obr_4group_non4bit_scheme', 'order2')).lower() in {"ternary", "ttb1"}:
                parts.append(f"dlt{getattr(args, 'svd_2bit_obr_ternary_delta_factor', 0.7):.2f}")
        if getattr(args, 'svd_2bit_obr_twogroup_enable', False):
            if getattr(args, 'svd_2bit_obr_twogroup_adaptive_enable', False):
                parts.append(
                    f"casc42adp_m{getattr(args, 'svd_2bit_obr_twogroup_adaptive_max_stages', 3)}"
                    f"_e{int(round(getattr(args, 'svd_2bit_obr_twogroup_adaptive_e128_threshold', 0.2) * 100)):02d}"
                    f"_b{int(round(getattr(args, 'svd_2bit_obr_twogroup_adaptive_base_ratio', 0.1) * 100)):02d}"
                )
                if bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio', False)):
                    parts.append(
                        f"fixr4{getattr(args, 'svd_2bit_obr_twogroup_4bit_ratio', 0.1):.2f}"
                    )
                if bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor', False)):
                    parts.append("lnk0")
            else:
                parts.append(
                    f"casc42_r4{getattr(args, 'svd_2bit_obr_twogroup_4bit_ratio', 0.1):.2f}"
                )
        if getattr(args, 'svd_obr_interblock_gptq_propagation_disable', False):
            parts.append("nogptqprop")
        if getattr(args, 'svd_2bit_extra_8bit_stage_enable', False):
            parts.append(f"plus8r{getattr(args, 'svd_2bit_extra_8bit_stage_rank', 64)}")
        if getattr(args, 'svd_binary_refit_enable', False):
            parts.append("binrefit")
            parts.append(f"ep{getattr(args, 'svd_binary_refit_epochs', 5)}")
        if getattr(args, 'svd_bit_stats_enable', False):
            parts.append("bitstats")
    
    if getattr(args, 'disable_salient_mask', False):
        parts.append("ns")  # no salient
    
    if getattr(args, 'row_wise_split', False):
        parts.append("rowwise")  # row-wise split
    
    if getattr(args, 'binary_residual', False):
        parts.append("binary")  # binary residual
    
    if getattr(args, 'smoothquant_enable', False):
        parts.append("smooth")  # smoothquant
    
    if getattr(args, 'salient_first_enable', False):
        parts.append("salient_first")  # salient-first mode
    
    if getattr(args, 'svd_one_opt', False):
        parts.append("1_opt")  # one-opt mode
    
    if getattr(args, 'svd_r_first', False):
        parts.append("r_first")  # R-first mode
        if getattr(args, 'svd_r_first_refit_l', False):
            parts.append("refitL")
    
    if getattr(args, 'svd_row_prune_search_enable', False) and getattr(args, 'svd_row_prune_clip_search_enable', False):
        parts.append("rpclip")
    if getattr(args, 'svd_row_prune_search_enable', False) and int(getattr(args, 'svd_row_prune_act_topk', 0)) > 0:
        parts.append(f"rpactk{int(getattr(args, 'svd_row_prune_act_topk', 0))}")
    if getattr(args, 'svd_row_prune_search_enable', False) and getattr(args, 'svd_row_prune_offline_refill_enable', False):
        parts.append("rprefill")
    if getattr(args, 'post_gptq_refill_enable', False):
        row_pct = int(round(float(getattr(args, 'post_gptq_refill_row_ratio', 0.2)) * 100))
        pos_pct = int(round(float(getattr(args, 'post_gptq_refill_pos_ratio', 0.2)) * 100))
        parts.append(f"postrefill_r{row_pct}_p{pos_pct}")
    if getattr(args, 'in_gptq_refill_enable', False):
        row_pct = int(round(float(getattr(args, 'in_gptq_refill_row_ratio', 0.2)) * 100))
        pos_pct = int(round(float(getattr(args, 'in_gptq_refill_pos_ratio', 0.2)) * 100))
        parts.append(f"ingptqrefill_r{row_pct}_p{pos_pct}")
    
    if getattr(args, 'bias_correction', False):
        parts.append("bias_corr")  # bias correction
    if getattr(args, 'structure_prune', False):
        parts.append("structure_prune")  # structure prune
        parts.append(f"prune_n{getattr(args, 'prune_n', 0)}")
        parts.append(f"prune_m{getattr(args, 'prune_m', 0)}")
    # parts.append("1254wu3bit")
    if getattr(args, 'dataset', 'wikitext2'):
        parts.append(getattr(args, 'dataset', 'wikitext2'))
    save_title = "_".join(parts)
    save_dir = "./output/" + save_title.replace("/", "_")
    
    model_loaded_from_save = False
    if os.path.exists(save_dir) and os.path.isdir(save_dir):
        print(f"Found saved model at {save_dir}, loading and evaluating...")
        model = get_model(save_dir)
        model.eval()
        model_loaded_from_save = True

        summary_path = os.path.join(save_dir, 'quantization_summary.json')
        if os.path.exists(summary_path):
            try:
                import json
                with open(summary_path, 'r') as f:
                    save_meta = json.load(f)
                qcfg = save_meta.get('billm_quantization_config', save_meta.get('quantization_config', {}))
                if isinstance(qcfg, dict):
                    logger.info(
                        "Saved quantization summary: "
                        f"residual_backend={qcfg.get('residual_backend', 'unknown')}, "
                        f"obr={qcfg.get('obr_enabled', 'unknown')}, "
                        f"gptq_enabled={qcfg.get('gptq_enabled', 'unknown')}, "
                        f"disable_gptq={qcfg.get('disable_gptq', 'unknown')}, "
                        f"interblock_gptq_propagation={qcfg.get('interblock_gptq_propagation_enabled', 'unknown')}, svd_obr_scheme={qcfg.get('svd_obr_scheme', 'none')}"
                    )
            except Exception as e:
                logger.warning(f"Failed to read saved quantization summary: {e}")
        
        try:
            from smooth_linear import SmoothLinear
            
            # 检查是否有正向 SmoothQuant 参数
            has_smooth = False
            for name, module in model.named_modules():
                if isinstance(module, (nn.Linear, transformers.Conv1D)) and hasattr(module, 'smooth_scale'):
                    smooth_scale = module.smooth_scale
                    if smooth_scale.numel() > 1:  # 不是默认值
                        has_smooth = True
                        break
            
            if has_smooth:
                logger.info("Found SmoothQuant parameters, replacing with SmoothLinear...")
                for name, module in model.named_modules():
                    if isinstance(module, (nn.Linear, transformers.Conv1D)) and hasattr(module, 'smooth_scale'):
                        smooth_scale = module.smooth_scale
                        if smooth_scale.numel() > 1:  # 不是默认值
                            # 获取权重和偏置
                            if isinstance(module, transformers.Conv1D):
                                weight = module.weight.t()
                                in_features, out_features = weight.shape[1], weight.shape[0]
                            else:
                                weight = module.weight
                                in_features, out_features = weight.shape[1], weight.shape[0]
                            
                            smooth_linear = SmoothLinear(
                                in_features=in_features,
                                out_features=out_features,
                                bias=module.bias is not None,
                                weight=weight,
                                smooth_scale=smooth_scale,
                                layer_name=name
                            )
                            
                            # 替换模块
                            parts = name.split('.')
                            parent = model
                            for part in parts[:-1]:
                                parent = getattr(parent, part)
                            setattr(parent, parts[-1], smooth_linear)
                
                logger.info("✓ Replaced Linear layers with SmoothLinear and loaded SmoothQuant parameters")
        except Exception as e:
            logger.warning(f"Warning: Failed to replace with SmoothLinear: {e}")
    
    else: 
        model = get_model(args.model)
        model.eval()
        
        tick = time.time()
        # ⭐ 对于长上下文模型，使用 2048 进行 calibration（避免 OOM）
        calibration_seqlen = min(model.seqlen, 2048)
        logger.info(f"Model max seqlen: {model.seqlen}, Using calibration seqlen: {calibration_seqlen}")
        logger.info(f"Dataset: {args.dataset}")
        dataloader, testloader = get_loaders(
            args.dataset,
            nsamples=args.nsamples,
            seed=args.seed,
            model=args.model,
            seqlen=calibration_seqlen,
        )
        quant_sequential(model, dataloader, device)
        logger.info(f"quantization time: {time.time() - tick} s")

    if args.save and not model_loaded_from_save:
        import json
        import sys
        from datetime import datetime

        model.save_pretrained(save_dir)
        logger.info(f"Model saved to: {save_dir}")

        obr_enabled = bool(getattr(args, 'obr', False))
        disable_gptq = bool(getattr(args, 'disable_gptq', False))
        gptq_enabled = not disable_gptq
        interblock_gptq_prop_enabled = not bool(
            getattr(args, 'svd_obr_interblock_gptq_propagation_disable', False)
        )

        if obr_enabled:
            residual_backend = "obr"
        elif gptq_enabled:
            residual_backend = "gptq"
        else:
            residual_backend = "rtn_no_gptq"

        quant_cfg = {
            "method": "billm",
            "low_quant_method": getattr(args, 'low_quant_method', None),
            "pure_gptq": False,
            "std_gptq_enable": bool(getattr(args, 'std_gptq_enable', False)),
            "std_gptq_static_groups": bool(getattr(args, 'std_gptq_static_groups', True)),
            "block_gptq_enable": bool(getattr(args, 'std_gptq_enable', False)),
            "block_gptq_static_groups": bool(getattr(args, 'std_gptq_static_groups', True)),
            "post_gptq_refill_enable": bool(getattr(args, 'post_gptq_refill_enable', False)),
            "post_gptq_refill_row_ratio": float(getattr(args, 'post_gptq_refill_row_ratio', 0.2)),
            "post_gptq_refill_pos_ratio": float(getattr(args, 'post_gptq_refill_pos_ratio', 0.2)),
            "in_gptq_refill_enable": bool(getattr(args, 'in_gptq_refill_enable', False)),
            "in_gptq_refill_row_ratio": float(getattr(args, 'in_gptq_refill_row_ratio', 0.2)),
            "in_gptq_refill_pos_ratio": float(getattr(args, 'in_gptq_refill_pos_ratio', 0.2)),
            "svd256_then_pure_gptq2bit": bool(getattr(args, 'svd256_then_pure_gptq2bit', False)),
            "svd256_then_pure_gptq3bit": bool(getattr(args, 'svd256_then_pure_gptq3bit', False)),
            "gptq_groupsize": int(getattr(args, 'gptq_groupsize', -1)),
            "gptq_static_groups": bool(getattr(args, 'gptq_static_groups', False)),
            "gptq_act_order": bool(getattr(args, 'gptq_act_order', False)),
            "gptq_true_sequential": bool(getattr(args, 'gptq_true_sequential', False)),
            "gptq_sym": bool(getattr(args, 'gptq_sym', False)),
            "gptq2bit_prune_search_enable": bool(getattr(args, 'gptq2bit_prune_search_enable', False)),
            "gptq2bit_prune_m_candidates": list(getattr(args, 'gptq2bit_prune_m_candidates', list(range(0, 97, 4)))),
            "gptq3bit_prune_search_enable": bool(getattr(args, 'gptq3bit_prune_search_enable', False)),
            "gptq3bit_prune_m_candidates": list(getattr(args, 'gptq3bit_prune_m_candidates', list(range(0, 97, 4)))),
            "residual_backend": residual_backend,
            "obr_enabled": obr_enabled,
            "gptq_enabled": gptq_enabled,
            "disable_gptq": disable_gptq,
            "interblock_gptq_propagation_enabled": interblock_gptq_prop_enabled,
            "svd_enable": bool(getattr(args, 'svd_enable', False)),
            "svd_auto_rank_deltae_stop": bool(getattr(args, 'svd_auto_rank_deltae_stop', False)),
            "svd_rank": int(getattr(args, 'svd_rank', 0)),
            "svd_num_iters": int(getattr(args, 'svd_num_iters', 1)),
            "svd_2bit_stages": int(getattr(args, 'svd_2bit_stages', 0)),
            "svd_lowrank_fp16": bool(getattr(args, 'svd_lowrank_fp16', False)),
            "svd_drop_residual": bool(getattr(args, 'svd_drop_residual', False)),
            "svd_2bit_obr_cascade_enable": bool(getattr(args, 'svd_2bit_obr_cascade_enable', False)),
            "svd_2bit_obr_4group_enable": bool(getattr(args, 'svd_2bit_obr_4group_enable', False)),
            "svd_2bit_obr_twogroup_enable": bool(getattr(args, 'svd_2bit_obr_twogroup_enable', False)),
            "svd_2bit_obr_twogroup_adaptive_enable": bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_enable', False)),
            "svd_2bit_obr_twogroup_adaptive_max_stages": int(getattr(args, 'svd_2bit_obr_twogroup_adaptive_max_stages', 3)),
            "svd_2bit_obr_twogroup_adaptive_e128_threshold": float(getattr(args, 'svd_2bit_obr_twogroup_adaptive_e128_threshold', 0.2)),
            "svd_2bit_obr_twogroup_adaptive_base_ratio": float(getattr(args, 'svd_2bit_obr_twogroup_adaptive_base_ratio', 0.1)),
            "svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio": bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_fixed_4bit_ratio', False)),
            "svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor": bool(getattr(args, 'svd_2bit_obr_twogroup_adaptive_linked_ratio_no_floor', False)),
            "svd_obr_scheme": str(getattr(args, 'svd_obr_scheme', 'none')),
            "svd_2bit_obr_twogroup_4bit_ratio": float(getattr(args, 'svd_2bit_obr_twogroup_4bit_ratio', 0.1)),
            "svd_2bit_obr_4group_4bit_ratio": float(getattr(args, 'svd_2bit_obr_4group_4bit_ratio', 0.1)),
            "svd_2bit_obr_4group_2bit_ratio_1": float(getattr(args, 'svd_2bit_obr_4group_2bit_ratio_1', 0.3)),
            "svd_2bit_obr_4group_2bit_ratio_2": float(getattr(args, 'svd_2bit_obr_4group_2bit_ratio_2', 0.3)),
            "svd_bit_stats_enable": bool(getattr(args, 'svd_bit_stats_enable', False)),
            "svd_bit_stats_scale_bits": float(getattr(args, 'svd_bit_stats_scale_bits', 16)),
            "svd_bit_stats_zp_bits": float(getattr(args, 'svd_bit_stats_zp_bits', 8)),
            "svd_bit_stats_binary_data_bits": float(getattr(args, 'svd_bit_stats_binary_data_bits', 2.0)),
            "svd_row_prune_search_enable": bool(getattr(args, 'svd_row_prune_search_enable', False)),
            "svd_row_prune_m_candidates": list(getattr(args, 'svd_row_prune_m_candidates', [0, 4, 8, 16, 32, 64])),
            "svd_row_prune_score_metric": str(getattr(args, 'svd_row_prune_score_metric', 'magnitude')),
            "svd_row_prune_binary_order": int(getattr(args, 'svd_row_prune_binary_order', 2)),
            "svd_row_prune_quant_scheme": str(getattr(args, 'svd_row_prune_quant_scheme', 'binary')),
            "svd_row_prune_clip_search_enable": bool(getattr(args, 'svd_row_prune_clip_search_enable', False)),
            "svd_row_prune_clip_candidates": list(getattr(args, 'svd_row_prune_clip_candidates', [1.0, 0.999, 0.995, 0.99, 0.98, 0.96])),
            "svd_row_prune_clip_min_value": float(getattr(args, 'svd_row_prune_clip_min_value', 1e-8)),
            "svd_row_prune_act_topk": int(getattr(args, 'svd_row_prune_act_topk', 3)),
            "svd_row_prune_offline_refill_enable": bool(getattr(args, 'svd_row_prune_offline_refill_enable', False)),
            "svd_row_prune_offline_refill_sweeps": int(getattr(args, 'svd_row_prune_offline_refill_sweeps', 1)),
            "svd_row_prune_offline_refill_max_positions": int(getattr(args, 'svd_row_prune_offline_refill_max_positions', -1)),
            "svd_row_prune_offline_refill_include_zero": bool(getattr(args, 'svd_row_prune_offline_refill_include_zero', True)),
            "svd_r_first": bool(getattr(args, 'svd_r_first', False)),
            "svd_r_first_refit_l": bool(getattr(args, 'svd_r_first_refit_l', False)),
        }

        summary_payload = {
            "created_at_utc": datetime.utcnow().isoformat() + "Z",
            "model": getattr(args, 'model', None),
            "dataset": getattr(args, 'dataset', None),
            "save_title": save_title,
            "save_dir": save_dir,
            "command": " ".join(sys.argv),
            "billm_quantization_config": quant_cfg,
        }

        summary_path = os.path.join(save_dir, 'quantization_summary.json')
        with open(summary_path, 'w') as f:
            json.dump(summary_payload, f, indent=2)
        logger.info(f"✓ Saved quantization summary to: {summary_path}")
        logger.info(
            "  Save summary: "
            f"residual_backend={residual_backend}, obr={obr_enabled}, "
            f"gptq_enabled={gptq_enabled}, disable_gptq={disable_gptq}, "
            f"interblock_gptq_propagation={interblock_gptq_prop_enabled}, svd_obr_scheme={getattr(args, 'svd_obr_scheme', 'none')}"
        )

        config_path = os.path.join(save_dir, 'config.json')
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                config = json.load(f)

            # NOTE: `quantization_config` is a reserved HF field.
            # Keep BiLLM metadata under billm_* keys only.
            legacy_qcfg = config.pop('quantization_config', None)
            if legacy_qcfg is not None:
                logger.info("Removed quantization_config from config.json to avoid HF loader conflicts")

            config['billm_quantization_config'] = quant_cfg
            config['billm_quantization_summary_file'] = 'quantization_summary.json'
            config['quantization_summary_file'] = 'quantization_summary.json'
            with open(config_path, 'w') as f:
                json.dump(config, f, indent=2)
            logger.info("✓ Added billm_quantization_config to config.json")
    for dataset in ["wikitext2", "ptb", "c4"]:
        try:
            eval_seqlen = min(model.seqlen, 2048)
            dataloader, testloader = get_loaders(
                dataset, seed=args.seed, seqlen=eval_seqlen, model=args.model
            )
            logger.info(dataset)
            model_str = args.model.lower()
            if "opt" in model_str:
                from eval_ppl_utils import opt_eval

                opt_eval(model, testloader, device, dataset, args.log_wandb)
            elif "llama" in model_str:
                from eval_ppl_utils import llama_eval

                llama_eval(model, testloader, device, dataset, args.log_wandb)
            elif "qwen" in model_str:
                from eval_ppl_utils import llama_eval

                llama_eval(model, testloader, device, dataset, args.log_wandb)
        except Exception as e:
            logger.error(f"Failed to evaluate on {dataset}: {e}")
            logger.warning(f"Skipping {dataset} evaluation and continuing with next dataset...")
            import traceback
            traceback.print_exc()
            continue
